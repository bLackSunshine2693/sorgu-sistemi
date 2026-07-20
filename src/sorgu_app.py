"""Ofis Sorgu Sistemi v4.0"""
import os,sys,hashlib,logging,socket,threading,webbrowser,datetime
from flask import Flask,request,jsonify,session

IS_FROZEN=getattr(sys,"frozen",False)
BASE_DIR=os.path.dirname(sys.executable) if IS_FROZEN else os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)
logging.basicConfig(level=logging.INFO,format="%(asctime)s %(levelname)-7s %(message)s",datefmt="%H:%M:%S")
log=logging.getLogger("sorgu")
if IS_FROZEN:
    import logging as _lg
    fh=_lg.FileHandler(os.path.join(BASE_DIR,"sorgu.log"),encoding="utf-8-sig")
    fh.setFormatter(_lg.Formatter("%(asctime)s %(levelname)-7s %(message)s","%H:%M:%S"))
    _lg.getLogger().addHandler(fh)

# ── LİSANS ───────────────────────────────────────────────────────────────────
try:
    from sorgu_license import check_license, ADMIN_PASS_HASH as LIC_ADMIN_HASH
    _ok,_msg,_lic=check_license(os.path.join(BASE_DIR,"license.lic"))
    log.info(f"Lisans: {_msg}")
    if not _ok:
        try:
            import tkinter as tk,tkinter.messagebox as mb
            r=tk.Tk();r.withdraw();mb.showerror("Lisans Hatası",f"{_msg}\n\nYöneticinizle iletişime geçin.");r.destroy()
        except: print(f"[HATA] {_msg}")
        sys.exit(1)
    _PLAN=_lic.get("plan","?");_BRANCH=_lic.get("branch","");_ENDDATE=_lic.get("end_date","");_ADMIN_MOD=_lic.get("admin_mod",False)
except ImportError:
    log.warning("sorgu_license.py bulunamadı — geliştirme modu")
    _PLAN="DEV";_BRANCH="";_ENDDATE=""

# ── ŞİFRELER ─────────────────────────────────────────────────────────────────
APP_PASSWORD_HASH="35ce5f0804be4660182f9df72f5b18f5f371d1d84565c6841acf608b57c2f608"

# ── VERSİYON & GÜNCELLEME ─────────────────────────────────────────────────────
CURRENT_VERSION   = "4.0.4"
GITHUB_RAW_URL    = "https://raw.githubusercontent.com/bLackSunshine2693/sorgu-version/main/version.json"
GITHUB_RELEASE_URL= "https://github.com/bLackSunshine2693/sorgu-sistemi/releases/latest/download/SorguSistemi.zip"
# ↑ GitHub repo adresinizi buraya yazın

def _db_pwd():
    k=b"SorguSistemi2024"
    e=bytes([61,87,58,14,84,123,32,23,14,46,55,28,117,5,22,7,40,38,0,70,64,37,92,35,80,23,37,72,22,90,101,0,52,8,9,17,25,99,34,95,77,0,88,61,65,83,76,117,32,28])
    return "".join(chr(e[i]^k[i%len(k)])for i in range(len(e)))

MARIADB={"host":"127.0.0.1","port":3307,"user":"root","password":_db_pwd()}
app=Flask(__name__)
app.secret_key="sorgu_2024_gizli_xK9"
app.config['JSON_SORT_KEYS']=False

# ── DB CONFIG — Sıralı ───────────────────────────────────────────────────────
DB_CONFIG={
    "tcpro": {"label":"Ad Soyad Sorgusu","icon":"👤","color":"#8b5cf6",
              "db":"tc","table":"tcpro","tc_col":"TC",
              "columns":["TC","AD","SOYAD","BABAADI","ANNEADI",
                         "DOGUMTARIHI","DOGUMYERI","MEMLEKETIL",
                         "MEMLEKETILCE","MEMLEKETKOY","ADRESIL","ADRESILCE"],
              "filters":[
                  {"key":"ADRESIL","label":"Adres İl","type":"text"},
                  {"key":"ADRESILCE","label":"Adres İlçe","type":"text"},
              ],
              "multi_search":True,"virtual":False},
    "gsm":   {"label":"GSM Bilgisi","icon":"📱","color":"#10b981",
              "virtual":True,"multi_modal":True,
              "modes":[
                  {"key":"tc2gsm",  "label":"TC → GSM"},
                  {"key":"gsm2tc",  "label":"GSM → TC"},
                  {"key":"tc2vergi","label":"TC → Vergi No"},
                  {"key":"vergi2tc","label":"Vergi No → TC"},
              ]},
    "adres": {"label":"Adres Bilgisi","icon":"📍","color":"#3b82f6",
              "db":"adres","table":"datam","tc_col":"KimlikNo",
              "columns":["KimlikNo","AdSoyad","DogumYeri","VergiNumarasi","Ikametgah"],
              "filters":[],"multi_search":False,"virtual":False},
    "komsular":{"label":"Aynı Hanede Oturanlar","icon":"🏘️","color":"#84cc16",
                "virtual":True,"mod":"adres"},
    "aile":  {"label":"Bizzat ve Aile Sorgusu","icon":"🧬","color":"#06b6d4","virtual":True},
    "tapu":  {"label":"Tapu Bilgisi","icon":"🏠","color":"#ef4444",
              "virtual":True,"multi_modal":True,
              "modes":[
                  {"key":"tc2tapu",    "label":"TC → Tapu"},
                  {"key":"parsel2kisi","label":"İl/İlçe/Ada/Parsel Sorgu"},
              ]},
    "sgk":   {"label":"İşyeri Sorgusu","icon":"🏢","color":"#f59e0b",
              "virtual":True,"multi_modal":True,
              "modes":[
                  {"key":"tc2isyeri", "label":"TC → İşyeri"},
                  {"key":"tc2calisanlar","label":"Aynı İşyerindekiler"},
              ]},
    "toplu": {"label":"Toplu GSM Sorgu","icon":"📋","color":"#e879f9",
              "virtual":True,"multi_modal":True,"admin_only":True,
              "modes":[
                  {"key":"bizzat",   "label":"Bizzat GSM"},
                  {"key":"aile",     "label":"Toplu Aile GSM"},
                  {"key":"aile2",    "label":"Toplu Aile GSM (2. Derece)"},
              ]},
}

# ── DB BAĞLANTI ───────────────────────────────────────────────────────────────
def get_conn(db):
    import mysql.connector
    return mysql.connector.connect(**MARIADB,database=db,charset="utf8mb4",use_unicode=True,connection_timeout=10)

def clean_row(r):
    return {k:(v.strftime("%d.%m.%Y %H:%M") if isinstance(v,(datetime.datetime,datetime.date)) else ("" if v is None else v))
            for k,v in r.items()}

def fetch_gsm(tc):
    if not tc or len(tc)!=11: return []
    try:
        c=get_conn("gsm");cur=c.cursor(dictionary=True)
        cur.execute("SELECT GSM FROM `145mgsm` WHERE TC=%s LIMIT 10",[tc])
        r=[x["GSM"] for x in cur.fetchall() if x.get("GSM")]
        cur.close();c.close();return r
    except: return []

def fetch_person(tc):
    if not tc or len(tc)!=11: return None
    c=get_conn("tc");cur=c.cursor(dictionary=True)
    cur.execute("SELECT * FROM tcpro WHERE TC=%s LIMIT 1",[tc])
    r=cur.fetchone();cur.close();c.close()
    if not r: return None
    p=clean_row(r);p["__gsm"]=fetch_gsm(tc);return p

def add_gsm(persons):
    if isinstance(persons,list):
        for p in persons:
            if p and "__gsm" not in p: p["__gsm"]=fetch_gsm(p.get("TC","") or p.get("calisanKimlikNo",""))
        return persons
    return persons

def find_es(tc,cocuklar):
    s=set()
    for c in cocuklar:
        b,a=c.get("BABATC",""),c.get("ANNETC","")
        if b==tc and a: s.add(a)
        elif a==tc and b: s.add(b)
    return list(s)

# ── ROUTES ───────────────────────────────────────────────────────────────────
@app.route("/")
def index(): return HTML_PAGE

@app.route("/api/login",methods=["POST"])
def api_login():
    d=request.get_json(force=True) or {}
    if hashlib.sha256(d.get("password","").encode()).hexdigest()==APP_PASSWORD_HASH:
        session["auth"]=True;log.info("Giriş OK");return jsonify({"ok":True})
    return jsonify({"ok":False}),401

@app.route("/api/logout",methods=["POST"])
def api_logout(): session.clear();return jsonify({"ok":True})


@app.route("/api/version")
def api_version():
    """Mevcut versiyon bilgisi."""
    return jsonify({"version": CURRENT_VERSION})

@app.route("/api/update/check")
def api_update_check():
    """GitHub'dan yeni sürüm kontrolü."""
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    try:
        import requests as req
        r = req.get(GITHUB_RAW_URL, timeout=5)
        data = r.json()
        remote = data.get("version","0.0.0")
        has_update = remote > CURRENT_VERSION
        return jsonify({
            "current": CURRENT_VERSION,
            "remote":  remote,
            "has_update": has_update,
            "notes":   data.get("notes",""),
            "date":    data.get("date",""),
        })
    except Exception as e:
        return jsonify({"current": CURRENT_VERSION, "has_update": False, "error": str(e)})

# Güncelleme indirme durumu
_update_progress = {"pct": 0, "status": "idle", "error": ""}


@app.route("/api/update/apply", methods=["POST"])
def api_update_apply():
    """update.bat dosyasını çalıştır ve uygulamayı kapat."""
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    bat = os.path.join(BASE_DIR, "update.bat")
    log.info(f"update.bat aranıyor: {bat}")
    if not os.path.exists(bat):
        # İndirme tamamlandıysa bat yeniden oluşturulmuş olabilir
        log.error(f"update.bat bulunamadı: {bat}")
        return jsonify({"error":f"update.bat bulunamadı: {bat}"}),404
    try:
        import subprocess as _sp
        _sp.Popen(f'cmd /c start "" "{bat}"', shell=True)
        log.info("update.bat çalıştırıldı")
        return jsonify({"ok":True})
    except Exception as e:
        log.error(f"apply hatası: {e}")
        return jsonify({"error":str(e)}),500


@app.route("/api/set_update")
def api_set_update():
    """Launcher'dan güncelleme bildirimi al."""
    global _update_info
    v = request.args.get("v","")
    notes = request.args.get("notes","")
    url = request.args.get("url","")
    if v:
        _update_info = {"version":v,"notes":notes,"url":url}
        log.info(f"[GÜNCELLEME] Yeni sürüm: {v} — {notes}")
    return jsonify({"ok":True})

_update_info = {}

@app.route("/api/update/info")
def api_update_info():
    return jsonify(_update_info)

@app.route("/api/update/progress")
def api_update_progress():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    return jsonify(_update_progress)

@app.route("/api/update/download", methods=["POST"])
def api_update_download():
    """Yeni sürümü arka planda indir, progress takip et."""
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    global _update_progress
    if _update_progress["status"] == "downloading":
        return jsonify({"error":"İndirme zaten devam ediyor"}),400

    def do_download():
        global _update_progress
        try:
            import requests as req, tempfile, urllib.request
            _update_progress = {"pct": 0, "status": "downloading", "error": ""}
            log.info("Güncelleme indiriliyor...")

            # download_url'yi version.json'dan al
            try:
                import json as _json
                vr = req.get(GITHUB_RAW_URL, timeout=5)
                dl_url = vr.json().get("download_url", GITHUB_RELEASE_URL)
            except:
                dl_url = GITHUB_RELEASE_URL

            # Önce dosya boyutunu öğren
            head = req.head(dl_url, timeout=10, allow_redirects=True)
            total = int(head.headers.get("content-length", 0))

            r = req.get(dl_url, timeout=120, stream=True)
            r.raise_for_status()
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
            downloaded = 0
            for chunk in r.iter_content(65536):
                tmp.write(chunk)
                downloaded += len(chunk)
                if total > 0:
                    _update_progress["pct"] = min(99, int(downloaded/total*100))
            tmp.close()

            # update.bat yaz
            exe_path = sys.executable if getattr(sys,"frozen",False) else os.path.join(BASE_DIR,"SorguSistemi.exe")
            exe_name = os.path.basename(exe_path)
            bat = os.path.join(BASE_DIR, "update.bat")
            tmp_extract = os.path.join(BASE_DIR, "_update_tmp")
            with open(bat,"w",encoding="utf-8") as f:
                f.write(f'''@echo off
echo Guncelleme basliyor, lutfen bekleyin...
timeout /t 3 /nobreak >nul

:: Calisان process'i durdur
taskkill /F /IM "{exe_name}" /T 2>nul
taskkill /F /IM "mysqld.exe" /T 2>nul
timeout /t 2 /nobreak >nul

:: Zip'i gecici klasore coz
if exist "{tmp_extract}" rmdir /s /q "{tmp_extract}"
mkdir "{tmp_extract}"
powershell -command "Expand-Archive -Path '{tmp.name}' -DestinationPath '{tmp_extract}' -Force"

:: Dosyalari kopyala
xcopy /Y /E /I "{tmp_extract}\*" "{BASE_DIR}\" >nul

:: Temizle
rmdir /s /q "{tmp_extract}" 2>nul
del "{tmp.name}" 2>nul

:: Yeniden baslat
start "" "{exe_path}"
del "%~f0"
''')
            _update_progress = {"pct": 100, "status": "done", "bat": bat, "error": ""}
            log.info("Güncelleme hazır")
        except Exception as e:
            _update_progress = {"pct": 0, "status": "error", "error": str(e)}
            log.error(f"Güncelleme hatası: {e}")

    import threading
    threading.Thread(target=do_download, daemon=True).start()
    return jsonify({"ok": True, "message": "İndirme başladı"})

@app.route("/api/config")
def api_config():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    import datetime as dt
    kalan=None
    if _ENDDATE and _ENDDATE!="SINIRSIZ":
        try: kalan=(datetime.date.fromisoformat(_ENDDATE)-datetime.date.today()).days
        except: pass
    return jsonify({
        "dbs_list":[{"key":k,"label":v["label"],"icon":v["icon"],"color":v["color"],
                  "filters":v.get("filters",[]),"multi_search":v.get("multi_search",False),
                  "virtual":v.get("virtual",False),"mod":v.get("mod",""),
                  "multi_modal":v.get("multi_modal",False),"modes":v.get("modes",[])}
               for k,v in DB_CONFIG.items()],
        "license":{"plan":_PLAN,"branch":_BRANCH,"end_date":_ENDDATE,"kalan":kalan,"admin_mod":_ADMIN_MOD},
    })

@app.route("/api/query",methods=["POST"])
def api_query():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    db_key=d.get("db","");tc=d.get("tc","").strip().replace(" ","")
    extra=d.get("filters",{});multi=d.get("multi",{})
    if db_key not in DB_CONFIG or DB_CONFIG[db_key].get("virtual"):
        return jsonify({"error":"Geçersiz DB"}),400
    cfg=DB_CONFIG[db_key]
    tc_son = d.get("tc_son","").strip()
    if cfg.get("multi_search") and not tc and not tc_son:
        return api_query_multi(cfg,multi,extra)
    # tc_son varsa veya tc varsa ana koda devam
    # TC: tam (11 hane) veya kısmi (prefix) arama
    if tc:
        tc_clean = tc.replace(" ","").replace("-","")
        if not tc_clean.isdigit():
            return jsonify({"error":"TC sadece rakam içerebilir"}),400
        tc = tc_clean
    if not tc and not (cfg.get("multi_search") and any(multi.values())):
        return jsonify({"error":"TC veya arama kriteri giriniz"}),400
    try:
        c=get_conn(cfg["db"]);cur=c.cursor(dictionary=True)
        where=[];params=[]
        tc_son = d.get("tc_son","").strip()
        log.info(f"  [query] db={db_key} tc={repr(tc)} tc_son={repr(tc_son)}")
        if tc and len(tc)==11:
            where.append(f'`{cfg["tc_col"]}` = %s');params.append(tc)
        elif tc and tc_son and len(tc_son)==2:
            wildcards = 11 - len(tc) - len(tc_son)
            if wildcards >= 0:
                pattern = f"{tc}{'_'*wildcards}{tc_son}"
                where.append(f'`{cfg["tc_col"]}` LIKE %s');params.append(pattern)
            else:
                where.append(f'`{cfg["tc_col"]}` LIKE %s');params.append(f"{tc}%")
        elif tc:
            where.append(f'`{cfg["tc_col"]}` LIKE %s');params.append(f"{tc}%")
        elif tc_son and len(tc_son)==2:
            # Son 2 hane — LIKE ile eşleşme
            where.append(f'`{cfg["tc_col"]}` LIKE %s');params.append(f"%{tc_son}")
        for f in cfg.get("filters",[]):
            v=extra.get(f["key"],"").strip()
            if v: where.append(f'`{f["key"]}` LIKE %s');params.append(f"{v}%")
        # Multi alanları da ekle (tcpro için AD, SOYAD vs)
        multi_fields={"AD":"AD","SOYAD":"SOYAD","BABAADI":"BABAADI","ANNEADI":"ANNEADI"}
        multi_count=0
        for key,col in multi_fields.items():
            mv=str(multi.get(key,"")).strip()
            if mv:
                where.append(f'`{col}` LIKE %s');params.append(f"{mv}%")
                multi_count+=1
        if not where:
            return jsonify({"error":"Arama kriteri giriniz"}),400
        # Limit: TC varsa 5000, yoksa 1000
        has_tc_where = any(cfg["tc_col"] in w for w in where)
        cols=", ".join(f'`{col}`' for col in cfg["columns"])
        sql=f'SELECT {cols} FROM {cfg["table"]} WHERE {" AND ".join(where)} LIMIT 50000'
        log.info(f"  [{db_key}] TC={tc}")
        cur.execute(sql,params)
        rows=[clean_row(r) for r in cur.fetchall()]
        cols_out=list(rows[0].keys()) if rows else cfg["columns"]
        cur.close();c.close()
        # Adres sorgusuna GSM ekle
        if db_key=="adres" and rows:
            for r in rows:
                gsm=fetch_gsm(r.get("KimlikNo",""))
                r["GSM"]=", ".join(gsm) if gsm else "—"
            if "GSM" not in cols_out: cols_out.append("GSM")
        return jsonify({"ok":True,"rows":rows,"columns":cols_out,"count":len(rows)})
    except Exception as e:
        log.error(f"[{db_key}] {e}");return jsonify({"error":str(e)}),500

def api_query_multi(cfg,multi,extra={}):
    fields={"AD":"AD","SOYAD":"SOYAD","BABAADI":"BABAADI","ANNEADI":"ANNEADI",
            "DOGUMTARIHI":"DOGUMTARIHI","DOGUMYERI":"DOGUMYERI",
            "MEMLEKETIL":"MEMLEKETIL","MEMLEKETILCE":"MEMLEKETILCE"}
    where=[];params=[]
    for key,col in fields.items():
        v=str(multi.get(key,"")).strip()
        if v: where.append(f'`{col}` LIKE %s');params.append(f"{v}%")
    # Filtreler (Adres İl/İlçe)
    for f in cfg.get("filters",[]):
        v=extra.get(f["key"],"").strip()
        if v: where.append(f'`{f["key"]}` LIKE %s');params.append(f"{v}%")
    if not where: return jsonify({"error":"En az bir arama kriteri giriniz"}),400
    try:
        c=get_conn(cfg["db"]);cur=c.cursor(dictionary=True)
        cols=", ".join(f'`{col}`' for col in cfg["columns"])
        cur.execute(f'SELECT {cols} FROM {cfg["table"]} WHERE {" AND ".join(where)} LIMIT 100',params)
        rows=[clean_row(r) for r in cur.fetchall()]
        cols_out=list(rows[0].keys()) if rows else cfg["columns"]
        cur.close();c.close()
        return jsonify({"ok":True,"rows":rows,"columns":cols_out,"count":len(rows)})
    except Exception as e:
        log.error(f"[multi] {e}");return jsonify({"error":str(e)}),500

@app.route("/api/gsm_query",methods=["POST"])
def api_gsm_query():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    mode=d.get("mode","tc2gsm");val=d.get("val","").strip().replace(" ","")
    if not val: return jsonify({"error":"Değer giriniz"}),400
    try:
        if mode=="tc2gsm":
            if len(val)!=11 or not val.isdigit(): return jsonify({"error":"11 haneli TC giriniz"}),400
            c=get_conn("gsm");cur=c.cursor(dictionary=True)
            cur.execute("SELECT TC,GSM FROM `145mgsm` WHERE TC=%s LIMIT 100",[val])
            rows=[clean_row(r) for r in cur.fetchall()]
            # İsim ekle
            if rows:
                p=fetch_person(val)
                ad=f"{p.get('AD','')} {p.get('SOYAD','')}".strip() if p else "—"
                for r in rows: r["AD_SOYAD"]=ad
            cols=["AD_SOYAD","TC","GSM"]
        elif mode=="gsm2tc":
            c=get_conn("gsm");cur=c.cursor(dictionary=True)
            cur.execute("SELECT TC,GSM FROM `145mgsm` WHERE GSM=%s LIMIT 100",[val])
            rows=[clean_row(r) for r in cur.fetchall()]
            # Her TC için isim ekle
            for r in rows:
                p=fetch_person(r.get("TC",""))
                r["AD_SOYAD"]=f"{p.get('AD','')} {p.get('SOYAD','')}".strip() if p else "—"
            cols=["AD_SOYAD","TC","GSM"]
        elif mode=="tc2vergi":
            if len(val)!=11 or not val.isdigit(): return jsonify({"error":"11 haneli TC giriniz"}),400
            c=get_conn("adres");cur=c.cursor(dictionary=True)
            cur.execute("SELECT KimlikNo,AdSoyad,VergiNumarasi FROM datam WHERE KimlikNo=%s LIMIT 10",[val])
            rows=[clean_row(r) for r in cur.fetchall()]
            cols=["KimlikNo","AdSoyad","VergiNumarasi"]
        elif mode=="vergi2tc":
            c=get_conn("adres");cur=c.cursor(dictionary=True)
            cur.execute("SELECT KimlikNo,AdSoyad,VergiNumarasi FROM datam WHERE VergiNumarasi=%s LIMIT 100",[val])
            rows=[clean_row(r) for r in cur.fetchall()]
            cols=["KimlikNo","AdSoyad","VergiNumarasi"]
        else:
            return jsonify({"error":"Geçersiz mod"}),400
        cur.close();c.close()
        return jsonify({"ok":True,"rows":rows,"columns":cols,"count":len(rows)})
    except Exception as e:
        log.error(f"[gsm/{mode}] {e}");return jsonify({"error":str(e)}),500

@app.route("/api/tapu_query",methods=["POST"])
def api_tapu_query():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    mode=d.get("mode","tc2tapu")
    try:
        if mode=="tc2tapu":
            tc=d.get("tc","").strip().replace(" ","")
            if not tc or len(tc)!=11 or not tc.isdigit():
                return jsonify({"error":"11 haneli TC giriniz"}),400
            c=get_conn("tapu");cur=c.cursor(dictionary=True)
            cols_sql=",".join(["`Identify`","`Name`","`Surname`","`BabaAdi`",
                               "`İlBilgisi`","`İlceBilgisi`","`MahalleBilgisi`",
                               "`ZeminTipBilgisi`","`AdaBilgisi`","`ParselBilgisi`",
                               "`YuzolcumBilgisi`","`AnaTasinmazNitelik`","`BlokBilgisi`",
                               "`BagimsizBolumNo`","`ArsaPay`","`ArsaPayda`",
                               "`BagimsizBolumNitelik`","`IstirakNo`",
                               "`HissePay`","`HissePayda`","`EdinmeSebebi`","`TapuDate`","`Yevmiye`"])
            cur.execute(f"SELECT {cols_sql} FROM venom_tapu WHERE `Identify`=%s LIMIT 5000",[tc])
            rows=[clean_row(r) for r in cur.fetchall()]
            # GSM ekle
            gsm=fetch_gsm(tc)
            gsm_str=", ".join(gsm) if gsm else "—"
            for r in rows: r["GSM"]=gsm_str
            cols_out=list(rows[0].keys()) if rows else []
        elif mode=="parsel2kisi":
            il=d.get("il","").strip();ilce=d.get("ilce","").strip()
            ada=d.get("ada","").strip();parsel=d.get("parsel","").strip()
            where=[];params=[]
            if il: where.append("`İlBilgisi`=%s");params.append(il)
            if ilce: where.append("`İlceBilgisi`=%s");params.append(ilce)
            if ada: where.append("`AdaBilgisi`=%s");params.append(ada)
            if parsel: where.append("`ParselBilgisi`=%s");params.append(parsel)
            if not where: return jsonify({"error":"En az bir kriter giriniz"}),400
            c=get_conn("tapu");cur=c.cursor(dictionary=True)
            cur.execute(f"SELECT `Identify`,`Name`,`Surname`,`BabaAdi`,`İlBilgisi`,`İlceBilgisi`,`MahalleBilgisi`,`AdaBilgisi`,`ParselBilgisi`,`HissePay`,`HissePayda`,`TapuDate` FROM venom_tapu WHERE {' AND '.join(where)} LIMIT 5000",params)
            rows=[clean_row(r) for r in cur.fetchall()]
            # GSM ekle
            for r in rows:
                gsm=fetch_gsm(r.get("Identify",""))
                r["GSM"]=", ".join(gsm) if gsm else "—"
            cols_out=list(rows[0].keys()) if rows else []
        else:
            return jsonify({"error":"Geçersiz mod"}),400
        cur.close();c.close()
        return jsonify({"ok":True,"rows":rows,"columns":cols_out,"count":len(rows)})
    except Exception as e:
        log.error(f"[tapu/{mode}] {e}");return jsonify({"error":str(e)}),500

@app.route("/api/sgk_query",methods=["POST"])
def api_sgk_query():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    mode=d.get("mode","tc2isyeri")
    tc=d.get("tc","").strip().replace(" ","")
    if not tc or len(tc)!=11 or not tc.isdigit():
        return jsonify({"error":"11 haneli TC giriniz"}),400
    try:
        c=get_conn("sgk");cur=c.cursor(dictionary=True)
        if mode=="tc2isyeri":
            cur.execute("SELECT calisanKimlikNo,calisanAdSoyad,isyeriUnvani,isyeriSgkSicilNo,isyeriSektoru,calismaDurumu,iseGirisTarihi FROM theos WHERE calisanKimlikNo=%s LIMIT 50",[tc])
            rows=[clean_row(r) for r in cur.fetchall()]
            cols_out=["calisanKimlikNo","calisanAdSoyad","isyeriUnvani","isyeriSgkSicilNo","isyeriSektoru","calismaDurumu","iseGirisTarihi"]
        elif mode=="tc2calisanlar":
            # Önce kişinin işyeri sicil no'sunu bul
            cur.execute("SELECT isyeriSgkSicilNo,isyeriUnvani FROM theos WHERE calisanKimlikNo=%s LIMIT 1",[tc])
            isyeri=cur.fetchone()
            if not isyeri: return jsonify({"ok":True,"rows":[],"columns":[],"count":0,"info":"Bu TC için işyeri kaydı bulunamadı"})
            sicil=isyeri["isyeriSgkSicilNo"];unvan=isyeri["isyeriUnvani"]
            cur.execute("SELECT calisanKimlikNo,calisanAdSoyad,calismaDurumu,iseGirisTarihi FROM theos WHERE isyeriSgkSicilNo=%s LIMIT 5000",[sicil])
            rows=[clean_row(r) for r in cur.fetchall()]
            # GSM ekle
            for r in rows:
                gsm=fetch_gsm(r.get("calisanKimlikNo",""))
                r["GSM"]=", ".join(gsm) if gsm else "—"
            cols_out=["calisanKimlikNo","calisanAdSoyad","calismaDurumu","iseGirisTarihi","GSM"]
            return jsonify({"ok":True,"rows":rows,"columns":cols_out,"count":len(rows),"info":f"İşyeri: {unvan} | Sicil: {sicil}"})
        else:
            return jsonify({"error":"Geçersiz mod"}),400
        cur.close();c.close()
        return jsonify({"ok":True,"rows":rows,"columns":cols_out,"count":len(rows)})
    except Exception as e:
        log.error(f"[sgk/{mode}] {e}");return jsonify({"error":str(e)}),500

@app.route("/api/komsular",methods=["POST"])
def api_komsular():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    tc=d.get("tc","").strip().replace(" ","")
    if not tc or len(tc)!=11 or not tc.isdigit():
        return jsonify({"error":"11 haneli TC giriniz"}),400
    try:
        c=get_conn("adres");cur=c.cursor(dictionary=True)
        cur.execute("SELECT Ikametgah FROM datam WHERE KimlikNo=%s LIMIT 1",[tc])
        row=cur.fetchone()
        if not row or not row.get("Ikametgah"):
            cur.close();c.close()
            return jsonify({"error":"Bu TC için adres kaydı bulunamadı"}),404
        adres=row["Ikametgah"]
        cur.execute("SELECT KimlikNo,AdSoyad,Ikametgah FROM datam WHERE Ikametgah=%s AND KimlikNo!=%s LIMIT 5000",[adres,tc])
        komsu_rows=cur.fetchall();cur.close();c.close()
        rows=[]
        for r in komsu_rows:
            ktc=r.get("KimlikNo","")
            ad_soyad="—"
            try:
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT AD,SOYAD FROM tcpro WHERE TC=%s LIMIT 1",[ktc])
                p=cur2.fetchone()
                if p: ad_soyad=f"{p.get('AD','')} {p.get('SOYAD','')}".strip()
                cur2.close();c2.close()
            except: pass
            gsm=fetch_gsm(ktc)
            rows.append({"AD SOYAD":ad_soyad,"KimlikNo":ktc,"Ikametgah":r.get("Ikametgah",""),"GSM":", ".join(gsm) if gsm else "—"})
        # Sorgulanan kişiyi de listeye ekle (başa)
        kisi_row_data = None
        try:
            c3=get_conn("adres");cur3=c3.cursor(dictionary=True)
            cur3.execute("SELECT KimlikNo,AdSoyad,Ikametgah FROM datam WHERE KimlikNo=%s LIMIT 1",[tc])
            kr=cur3.fetchone();cur3.close();c3.close()
            if kr:
                p2=fetch_person(tc)
                ad2=f"{p2.get('AD','')} {p2.get('SOYAD','')}".strip() if p2 else kr.get("AdSoyad","—")
                gsm2=fetch_gsm(tc)
                kisi_row_data={"AD SOYAD":f"★ {ad2} (Sorgulanan)","KimlikNo":tc,
                               "Ikametgah":kr.get("Ikametgah",""),"GSM":", ".join(gsm2) if gsm2 else "—"}
        except: pass
        if kisi_row_data: rows=[kisi_row_data]+rows
        log.info(f"  [komsular] TC={tc} → {len(rows)} kişi")
        return jsonify({"ok":True,"rows":rows,"columns":["AD SOYAD","KimlikNo","Ikametgah","GSM"],"count":len(rows),"adres":adres})
    except Exception as e:
        log.error(f"[komsular] {e}");return jsonify({"error":str(e)}),500

@app.route("/api/aile",methods=["POST"])
def api_aile():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    tc=d.get("tc","").strip().replace(" ","")
    degree=int(d.get("degree",1))
    if not tc or len(tc)!=11 or not tc.isdigit():
        return jsonify({"error":"11 haneli TC giriniz"}),400
    try:
        result=build_aile(tc)
        # Derece 2+: Torunlar
        if degree>=2:
            torunlar=[]
            for cocuk in result.get("derece1",{}).get("cocuklar",[]):
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 20",
                    [cocuk.get("TC",""),cocuk.get("TC","")])
                t=[clean_row(r) for r in cur2.fetchall()]
                cur2.close();c2.close()
                add_gsm(t);torunlar.extend(t)
            result["torunlar"]=torunlar

        # Derece 3+: Yeğenler, Büyük torun
        if degree>=3:
            yegenler=[]
            for kard in result.get("derece1",{}).get("kardesler",[]):
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 20",
                    [kard.get("TC",""),kard.get("TC","")])
                y=[clean_row(r) for r in cur2.fetchall()]
                cur2.close();c2.close()
                add_gsm(y);yegenler.extend(y)
            result["yegenler"]=yegenler
            # Büyük torun (torunların çocukları)
            buyuk_torun=[]
            for tor in result.get("torunlar",[]):
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 10",
                    [tor.get("TC",""),tor.get("TC","")])
                bt=[clean_row(r) for r in cur2.fetchall()]
                cur2.close();c2.close()
                add_gsm(bt);buyuk_torun.extend(bt)
            result["buyuk_torun"]=buyuk_torun

        # Derece 4+: Kuzenler, B. Amca/dayı/hala/teyze, Büyük yeğen
        if degree>=4:
            kuzenler=[]
            d2=result.get("derece2",{})
            for rel in list(d2.get("amca_hala",[]))+list(d2.get("dayi_teyze",[])):
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 20",
                    [rel.get("TC",""),rel.get("TC","")])
                k=[clean_row(r) for r in cur2.fetchall()]
                cur2.close();c2.close()
                add_gsm(k);kuzenler.extend(k)
            result["kuzenler"]=kuzenler
            # B. Amca/hala/dayı/teyze (dede/nine'nin kardeşleri)
            buyuk_amca=[]
            # baba_baba → B. Amca(E)/B. Hala(K)
            # baba_anne → B. Amca(E)/B. Hala(K)
            # anne_baba → B. Dayı(E)/B. Teyze(K)
            # anne_anne → B. Dayı(E)/B. Teyze(K)
            taraf_map={"baba_baba":("B. Amca","B. Hala"),
                       "baba_anne":("B. Amca","B. Hala"),
                       "anne_baba":("B. Dayı","B. Teyze"),
                       "anne_anne":("B. Dayı","B. Teyze")}
            for dede_key,lbl_pair in taraf_map.items():
                dede=d2.get(dede_key)
                if not dede: continue
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                baba_tc2=dede.get("BABATC","")
                if baba_tc2:
                    cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 15",
                        [baba_tc2,dede.get("TC","")])
                    for r in cur2.fetchall():
                        p=clean_row(r)
                        cin=p.get("CINSIYET","")
                        p["__rol"]=lbl_pair[0] if cin=="E" else lbl_pair[1] if cin=="K" else f"{lbl_pair[0]}/{lbl_pair[1]}"
                        buyuk_amca.append(p)
                cur2.close();c2.close()
            add_gsm(buyuk_amca)
            result["buyuk_amca"]=buyuk_amca
            # Büyük yeğen (yeğenlerin çocukları)
            buyuk_yegen=[]
            for yeg in result.get("yegenler",[]):
                c2=get_conn("tc");cur2=c2.cursor(dictionary=True)
                cur2.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 10",
                    [yeg.get("TC",""),yeg.get("TC","")])
                by=[clean_row(r) for r in cur2.fetchall()]
                cur2.close();c2.close()
                add_gsm(by);buyuk_yegen.extend(by)
            result["buyuk_yegen"]=buyuk_yegen

        return jsonify({"ok":True,"degree":degree,**result})
    except Exception as e: log.error(f"[aile] {e}");return jsonify({"error":str(e)}),500

def build_aile(tc):
    p=fetch_person(tc)
    if not p: return {"kisi":None,"derece1":{},"derece2":{},"derece3":{},"es_tarafi":{}}
    baba_tc=p.get("BABATC","");anne_tc=p.get("ANNETC","")
    baba=fetch_person(baba_tc) if baba_tc else None
    anne=fetch_person(anne_tc) if anne_tc else None
    kardesler=[]
    if baba_tc:
        c=get_conn("tc");cur=c.cursor(dictionary=True)
        cur.execute("SELECT * FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 30",[baba_tc,tc])
        kardesler=add_gsm([clean_row(r) for r in cur.fetchall()]);cur.close();c.close()
    c=get_conn("tc");cur=c.cursor(dictionary=True)
    cur.execute("SELECT * FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 30",[tc,tc])
    cocuklar=add_gsm([clean_row(r) for r in cur.fetchall()]);cur.close();c.close()
    es_tariflari={}
    for es_tc in find_es(tc,cocuklar):
        es=fetch_person(es_tc)
        if not es: continue
        kb_tc=es.get("BABATC","");ka_tc=es.get("ANNETC","")
        kb=fetch_person(kb_tc) if kb_tc else None;ka=fetch_person(ka_tc) if ka_tc else None
        esk=[]
        if kb_tc:
            c=get_conn("tc");cur=c.cursor(dictionary=True)
            cur.execute("SELECT * FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 20",[kb_tc,es_tc])
            esk=add_gsm([clean_row(r) for r in cur.fetchall()]);cur.close();c.close()
        es_tariflari[es_tc]={"es":es,"kayin_baba":kb,"kayin_anne":ka,"es_kardesler":esk,
            "kayin_dede_b":fetch_person(kb.get("BABATC","")) if kb else None,
            "kayin_nine_b":fetch_person(kb.get("ANNETC","")) if kb else None,
            "kayin_dede_a":fetch_person(ka.get("BABATC","")) if ka else None,
            "kayin_nine_a":fetch_person(ka.get("ANNETC","")) if ka else None}
    bb=fetch_person(baba.get("BABATC","")) if baba else None
    ba=fetch_person(baba.get("ANNETC","")) if baba else None
    ab=fetch_person(anne.get("BABATC","")) if anne else None
    aa=fetch_person(anne.get("ANNETC","")) if anne else None
    ah=[]
    if baba and baba.get("BABATC"):
        c=get_conn("tc");cur=c.cursor(dictionary=True)
        cur.execute("SELECT * FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 20",[baba["BABATC"],baba_tc])
        ah=add_gsm([clean_row(r) for r in cur.fetchall()]);cur.close();c.close()
        for p2 in ah:
            cin=p2.get("CINSIYET","")
            p2["__rol"]="Amca" if cin=="E" else "Hala" if cin=="K" else "Amca/Hala"
    dt=[]
    if anne and anne.get("BABATC"):
        c=get_conn("tc");cur=c.cursor(dictionary=True)
        cur.execute("SELECT * FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 20",[anne["BABATC"],anne_tc])
        dt=add_gsm([clean_row(r) for r in cur.fetchall()]);cur.close();c.close()
        for p2 in dt:
            cin=p2.get("CINSIYET","")
            p2["__rol"]="Dayı" if cin=="E" else "Teyze" if cin=="K" else "Dayı/Teyze"
    return {"kisi":p,
        "derece1":{"baba":baba,"anne":anne,"kardesler":kardesler,"cocuklar":cocuklar},
        "derece2":{"baba_baba":bb,"baba_anne":ba,"anne_baba":ab,"anne_anne":aa,"amca_hala":ah,"dayi_teyze":dt},
        "derece3":{"bb_b":fetch_person(bb.get("BABATC","")) if bb else None,
                   "bb_a":fetch_person(bb.get("ANNETC","")) if bb else None,
                   "ba_b":fetch_person(ba.get("BABATC","")) if ba else None,
                   "ba_a":fetch_person(ba.get("ANNETC","")) if ba else None,
                   "ab_b":fetch_person(ab.get("BABATC","")) if ab else None,
                   "ab_a":fetch_person(ab.get("ANNETC","")) if ab else None,
                   "aa_b":fetch_person(aa.get("BABATC","")) if aa else None,
                   "aa_a":fetch_person(aa.get("ANNETC","")) if aa else None},
        "es_tarafi":es_tariflari}



@app.route("/api/toplu_gsm_preview", methods=["POST"])
def api_toplu_gsm_preview():
    """Toplu sorgu sonuçlarını ekranda göstermek için JSON döner."""
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d=request.get_json(force=True) or {}
    mode=d.get("mode","bizzat")
    tc_list=[str(t).strip() for t in d.get("tc_list",[]) if str(t).strip()]
    if not tc_list: return jsonify({"error":"TC listesi boş"}),400
    # TC limiti yok

    if mode=="bizzat":
        rows=[];cols=["TC","AD SOYAD","GSM"]
        for tc in tc_list:
            p=fetch_person(tc)
            name=f"{p.get('AD','')} {p.get('SOYAD','')}".strip() if p else "—"
            gsm_list=fetch_gsm(tc)
            if not gsm_list: rows.append({"TC":tc,"AD SOYAD":name,"GSM":"—"})
            else:
                for gsm in gsm_list: rows.append({"TC":tc,"AD SOYAD":name,"GSM":gsm})
        return jsonify({"ok":True,"rows":rows,"columns":cols,"count":len(tc_list)})

    elif mode=="aile":  # aile - hızlı mod: sadece 1. derece aile
        cols=["Sorgulanan TC","Yakınlık","Ad Soyad","TC","GSM"]
        rows=[]
        for i,tc in enumerate(tc_list):
            log.info(f"  [toplu/preview] TC={tc} ({i+1}/{len(tc_list)})")
            try:
                # Kişiyi al
                c=get_conn("tc");cur=c.cursor(dictionary=True)
                cur.execute("SELECT * FROM tcpro WHERE TC=%s LIMIT 1",[tc])
                kisi=cur.fetchone()
                if not kisi: cur.close();c.close();continue
                kisi=clean_row(kisi)

                # 1. derece sorgu (tek query)
                baba_tc=kisi.get("BABATC","");anne_tc=kisi.get("ANNETC","")
                aile_rows=[]

                # Bizzat
                gsm=fetch_gsm(tc)
                ad=f"{kisi.get('AD','')} {kisi.get('SOYAD','')}".strip()
                aile_rows.append({"Sorgulanan TC":tc,"Yakınlık":"bizzat","Ad Soyad":ad,"TC":tc,"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                # Baba
                if baba_tc:
                    cur.execute("SELECT * FROM tcpro WHERE TC=%s LIMIT 1",[baba_tc])
                    b=cur.fetchone()
                    if b:
                        b=clean_row(b);gsm=fetch_gsm(baba_tc)
                        ad=f"{b.get('AD','')} {b.get('SOYAD','')}".strip()
                        aile_rows.append({"Sorgulanan TC":tc,"Yakınlık":"Baba","Ad Soyad":ad,"TC":baba_tc,"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                # Anne
                if anne_tc:
                    cur.execute("SELECT * FROM tcpro WHERE TC=%s LIMIT 1",[anne_tc])
                    a=cur.fetchone()
                    if a:
                        a=clean_row(a);gsm=fetch_gsm(anne_tc)
                        ad=f"{a.get('AD','')} {a.get('SOYAD','')}".strip()
                        aile_rows.append({"Sorgulanan TC":tc,"Yakınlık":"Anne","Ad Soyad":ad,"TC":anne_tc,"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                # Kardeşler (baba TC üzerinden)
                if baba_tc:
                    cur.execute("SELECT TC,AD,SOYAD FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 20",[baba_tc,tc])
                    for k in cur.fetchall():
                        k=clean_row(k);gsm=fetch_gsm(k["TC"])
                        ad=f"{k.get('AD','')} {k.get('SOYAD','')}".strip()
                        aile_rows.append({"Sorgulanan TC":tc,"Yakınlık":"Kardeş","Ad Soyad":ad,"TC":k["TC"],"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                # Çocuklar
                cur.execute("SELECT TC,AD,SOYAD FROM tcpro WHERE BABATC=%s OR ANNETC=%s LIMIT 20",[tc,tc])
                for c2 in cur.fetchall():
                    c2=clean_row(c2);gsm=fetch_gsm(c2["TC"])
                    ad=f"{c2.get('AD','')} {c2.get('SOYAD','')}".strip()
                    aile_rows.append({"Sorgulanan TC":tc,"Yakınlık":"Çocuk","Ad Soyad":ad,"TC":c2["TC"],"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                cur.close();c.close()
                rows.extend(aile_rows)
                if i<len(tc_list)-1: rows.append({"__sep":True})
            except Exception as e:
                log.error(f"[toplu/aile] TC={tc} hata: {e}")
                continue

        return jsonify({"ok":True,"rows":rows,"columns":cols,"count":len(tc_list)})

    elif mode=="aile2":  # Sadece 2. derece
        cols=["Sorgulanan TC","Yakınlık","Ad Soyad","TC","GSM"]
        rows=[]
        for i,tc in enumerate(tc_list):
            log.info(f"  [toplu/aile2] TC={tc} ({i+1}/{len(tc_list)})")
            try:
                c=get_conn("tc");cur=c.cursor(dictionary=True)
                cur.execute("SELECT * FROM tcpro WHERE TC=%s LIMIT 1",[tc])
                kisi=cur.fetchone()
                if not kisi: cur.close();c.close();continue
                kisi=clean_row(kisi)
                baba_tc=kisi.get("BABATC","");anne_tc=kisi.get("ANNETC","")

                def add2(tc2,yak):
                    if not tc2: return None
                    cur.execute("SELECT TC,AD,SOYAD,BABATC,ANNETC FROM tcpro WHERE TC=%s LIMIT 1",[tc2])
                    p=cur.fetchone()
                    if not p: return None
                    p=clean_row(p);gsm=fetch_gsm(tc2)
                    ad=f"{p.get('AD','')} {p.get('SOYAD','')}".strip()
                    rows.append({"Sorgulanan TC":tc,"Yakınlık":yak,"Ad Soyad":ad,"TC":tc2,"GSM":", ".join(gsm) if gsm else "—","__sep":False})
                    return p

                # Bizzat
                gsm=fetch_gsm(tc)
                ad=f"{kisi.get('AD','')} {kisi.get('SOYAD','')}".strip()
                rows.append({"Sorgulanan TC":tc,"Yakınlık":"bizzat","Ad Soyad":ad,"TC":tc,"GSM":", ".join(gsm) if gsm else "—","__sep":False})

                # Baba/Anne bilgileri (listeye ekleme, 2. derece için kullan)
                baba=None;anne=None
                if baba_tc:
                    cur.execute("SELECT TC,AD,SOYAD,BABATC,ANNETC FROM tcpro WHERE TC=%s LIMIT 1",[baba_tc])
                    r=cur.fetchone()
                    if r: baba=clean_row(r)
                if anne_tc:
                    cur.execute("SELECT TC,AD,SOYAD,BABATC,ANNETC FROM tcpro WHERE TC=%s LIMIT 1",[anne_tc])
                    r=cur.fetchone()
                    if r: anne=clean_row(r)

                # Sadece 2. derece
                if baba:
                    add2(baba.get("BABATC",""),"Dede (B)")
                    add2(baba.get("ANNETC",""),"Nine (B)")
                    if baba.get("BABATC",""):
                        cur.execute("SELECT TC FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 15",[baba["BABATC"],baba_tc])
                        for a in cur.fetchall(): add2(a["TC"],"Amca/Hala")
                if anne:
                    add2(anne.get("BABATC",""),"Dede (A)")
                    add2(anne.get("ANNETC",""),"Nine (A)")
                    if anne.get("BABATC",""):
                        cur.execute("SELECT TC FROM tcpro WHERE BABATC=%s AND TC!=%s LIMIT 15",[anne["BABATC"],anne_tc])
                        for a in cur.fetchall(): add2(a["TC"],"Dayı/Teyze")

                cur.close();c.close()
                if i<len(tc_list)-1: rows.append({"__sep":True})
            except Exception as e:
                log.error(f"[toplu/aile2] TC={tc} hata: {e}"); continue

        return jsonify({"ok":True,"rows":rows,"columns":cols,"count":len(tc_list)})

@app.route("/api/toplu_gsm_excel", methods=["POST"])
def api_toplu_gsm_excel():
    """Toplu TC listesinden GSM Excel dosyası üretir."""
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    d = request.get_json(force=True) or {}
    mode = d.get("mode","bizzat")
    tc_list = [str(t).strip() for t in d.get("tc_list",[]) if str(t).strip()]
    if not tc_list: return jsonify({"error":"TC listesi boş"}),400

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file

        wb = openpyxl.Workbook()
        ws = wb.active
        BLACK = PatternFill(start_color="000000",end_color="000000",fill_type="solid")
        HDR   = PatternFill(start_color="1e2a40",end_color="1e2a40",fill_type="solid")
        HDR_F = Font(color="8fa3c0",bold=True,size=9)
        WHT_F = Font(color="FFFFFF",bold=True)

        def hdr_row(cells, fill=HDR, font=HDR_F):
            ws.append(cells)
            row = ws.max_row
            for col in range(1, len(cells)+1):
                c = ws.cell(row=row, column=col)
                c.fill = fill; c.font = font
                c.alignment = Alignment(horizontal="center",vertical="center")

        def sep_row(ncols):
            ws.append([""]*ncols)
            row = ws.max_row
            for col in range(1, ncols+1):
                ws.cell(row=row, column=col).fill = BLACK

        if mode == "bizzat":
            hdr_row(["TC","AD SOYAD","GSM"])
            for tc in tc_list:
                p = fetch_person(tc)
                name = f"{p.get('AD','')} {p.get('SOYAD','')}".strip() if p else "—"
                gsm_list = fetch_gsm(tc)
                if not gsm_list:
                    ws.append([tc, name, "—"])
                else:
                    for gsm in gsm_list:
                        ws.append([tc, name, gsm])
            # Kolon genişlikleri
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 14

        elif mode in ("aile","aile2"):  # 1. ve 2. derece aile
            NCOLS = 8
            hdr_row(["Sorgulanan TC","Yakınlık","Ad Soyad","TC",
                     "Doğum","Ölüm","İl","Cinsiyet","Medeni",
                     "GSM","GSM","GSM","GSM"])

            def p_row(p, yakinlik, sorgu_tc):
                if not p: return None
                gsm_list = p.get("__gsm",[])
                row = [sorgu_tc, yakinlik,
                       f"{p.get('AD','')} {p.get('SOYAD','')}".strip(),
                       p.get("TC",""), p.get("DOGUMTARIHI",""),
                       p.get("OLUMTARIHI","") or "YOK",
                       p.get("MEMLEKETIL",""), p.get("CINSIYET",""),
                       p.get("MEDENIHAL","")]
                for j in range(4):
                    row.append(gsm_list[j] if j < len(gsm_list) else "")
                return row

            for i, tc in enumerate(tc_list):
                log.info(f"  [toplu/aile] TC={tc} ({i+1}/{len(tc_list)})")
                aile = build_aile(tc)
                if not aile.get("kisi"): continue
                kisi=aile["kisi"]
                d1=aile.get("derece1",{})
                d2=aile.get("derece2",{})
                es=aile.get("es_tarafi",{})

                rows = []
                r = p_row(kisi,"bizzat",tc)
                if r: rows.append(r)
                if d1.get("baba"): rows.append(p_row(d1["baba"],"Baba",tc))
                if d1.get("anne"): rows.append(p_row(d1["anne"],"Anne",tc))
                for k in d1.get("kardesler",[]): rows.append(p_row(k,"Kardeş",tc))
                for c in d1.get("cocuklar",[]): rows.append(p_row(c,"Çocuk",tc))
                for et in es.values():
                    if et.get("es"): rows.append(p_row(et["es"],"Eş",tc))
                    if et.get("kayin_baba"): rows.append(p_row(et["kayin_baba"],"Kayınpeder",tc))
                    if et.get("kayin_anne"): rows.append(p_row(et["kayin_anne"],"Kayınvalide",tc))
                    for k in et.get("es_kardesler",[]): rows.append(p_row(k,"Kayın",tc))
                for key,lbl in [("baba_baba","Dede(B)"),("baba_anne","Nine(B)"),
                                 ("anne_baba","Dede(A)"),("anne_anne","Nine(A)")]:
                    if d2.get(key): rows.append(p_row(d2[key],lbl,tc))
                for a in d2.get("amca_hala",[]): rows.append(p_row(a,"Amca/Hala",tc))
                for a in d2.get("dayi_teyze",[]): rows.append(p_row(a,"Dayı/Teyze",tc))

                for row in rows:
                    if row: ws.append(row)

                # Siyah ayırıcı (son TC değilse)
                if i < len(tc_list)-1:
                    sep_row(NCOLS)

            # Kolon genişlikleri
            widths = [14,12,22,14,14,14,14,14]
            for i,w in enumerate(widths,1):
                ws.column_dimensions[get_column_letter(i)].width = w

        output = BytesIO()
        wb.save(output); output.seek(0)
        fname = f"Toplu_GSM_{mode}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name=fname)
    except ImportError:
        return jsonify({"error":"openpyxl kurulu değil: pip install openpyxl"}),500
    except Exception as e:
        log.error(f"[toplu_gsm] {e}"); return jsonify({"error":str(e)}),500

@app.route("/api/test_db")
def api_test_db():
    if not session.get("auth"): return jsonify({"error":"Giriş gerekli"}),401
    results={}
    test_map={"gsm":("gsm","`145mgsm`"),"adres":("adres","datam"),
              "tcpro":("tc","tcpro"),"sgk":("sgk","theos"),"tapu":("tapu","venom_tapu")}
    for k,(db,tbl) in test_map.items():
        try:
            c=get_conn(db);cur=c.cursor()
            cur.execute(f"SELECT COUNT(*) FROM {tbl}")
            cnt=cur.fetchone()[0];cur.close();c.close()
            results[k]={"ok":True,"count":cnt}
        except Exception as e:
            results[k]={"ok":False,"error":str(e)}
    return jsonify(results)


HTML_PAGE=r"""<!DOCTYPE html>
<html lang="tr" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sorgu Sistemi v4</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#0a0e1a;--s1:#0f1524;--s2:#141b2d;--s3:#1a2236;--b1:#1e2a40;--b2:#253047;
  --t1:#e8f0fe;--t2:#8fa3c0;--t3:#506080;
  --ac:#4f9eff;--gr:#00e5a0;--yw:#ffd166;--rd:#ff4f6b;--pr:#c084fc;--cy:#22d3ee;--pk:#f472b6;
  --shadow:rgba(0,0,0,.4);--glow-ac:rgba(79,158,255,.12);--font:'Inter','Segoe UI',system-ui,sans-serif}
[data-theme="light"]{
  --bg:#f0f4ff;--s1:#ffffff;--s2:#f8faff;--s3:#eef2ff;--b1:#dde3f0;--b2:#c8d2e8;
  --t1:#1a2236;--t2:#4a5880;--t3:#8090b0;
  --ac:#2563eb;--gr:#059669;--yw:#d97706;--rd:#dc2626;--pr:#7c3aed;--cy:#0891b2;--pk:#db2777;
  --shadow:rgba(0,0,50,.1);--glow-ac:rgba(37,99,235,.08)}
body{background:var(--bg);color:var(--t1);font-family:var(--font);height:100vh;overflow:hidden;display:flex;flex-direction:column}

/* LOGIN */
#ls{position:fixed;inset:0;display:flex;align-items:center;justify-content:center;background:var(--bg);z-index:1000}
.lc{background:var(--s1);border:1px solid var(--b1);border-radius:20px;padding:40px;width:360px;box-shadow:0 20px 60px var(--shadow)}
.lc-top{text-align:center;margin-bottom:24px}
.lc-icon{font-size:2.5rem;margin-bottom:8px}
.lc h2{font-size:1.2rem;font-weight:700}
.lc p{color:var(--t2);font-size:.85rem;margin-top:4px}
.lf{margin:20px 0 16px}
.lf label{display:block;font-size:.72rem;font-weight:600;color:var(--t2);margin-bottom:6px;text-transform:uppercase;letter-spacing:.07em}
.lf input{width:100%;padding:11px 14px;background:var(--s2);border:1.5px solid var(--b2);border-radius:10px;color:var(--t1);font-size:.95rem;outline:none;font-family:var(--font)}
.lf input:focus{border-color:var(--ac)}
.lbtn{width:100%;padding:12px;background:var(--ac);border:none;border-radius:10px;color:#fff;font-size:.92rem;font-weight:700;cursor:pointer;font-family:var(--font)}
.lbtn:hover{opacity:.9}
.lerr{margin-top:10px;padding:9px 14px;background:rgba(255,79,107,.1);border:1px solid rgba(255,79,107,.3);border-radius:8px;color:var(--rd);font-size:.82rem;text-align:center;display:none}

/* LAYOUT */
#app{display:none;height:100vh;flex-direction:row}
.sidebar{width:220px;background:var(--s1);border-right:1px solid var(--b1);display:flex;flex-direction:column;flex-shrink:0;overflow:hidden}
.sb-header{padding:14px 12px 10px;border-bottom:1px solid var(--b1)}
.sb-logo{font-size:.9rem;font-weight:800;display:flex;align-items:center;gap:8px}
.sb-logo-icon{width:28px;height:28px;border-radius:7px;background:var(--ac);display:flex;align-items:center;justify-content:center;font-size:.85rem;flex-shrink:0}
.sb-sub{font-size:.65rem;color:var(--t2);margin-top:2px}
.sb-nav{padding:8px 6px;flex:1;overflow-y:auto}
.sb-sec{font-size:.6rem;font-weight:700;color:var(--t3);letter-spacing:.1em;text-transform:uppercase;padding:6px 6px 3px}
.nav-item{display:flex;align-items:center;gap:8px;padding:8px 8px;border-radius:9px;cursor:pointer;border:1px solid transparent;margin-bottom:2px;user-select:none;transition:all .15s}
.nav-item:hover{background:var(--s3);border-color:var(--b1)}
.nav-item.active{background:var(--glow-ac);border-color:rgba(79,158,255,.2);color:var(--ac)}
.nav-icon{width:28px;height:28px;border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:.85rem;flex-shrink:0;background:var(--s3);border:1px solid var(--b1)}
.nav-label{font-size:.78rem;font-weight:500}
.sb-lic{padding:8px 10px 6px;border-top:1px solid var(--b1);flex-shrink:0}
.sb-lic-badge{font-size:.67rem;font-weight:700;padding:2px 8px;border-radius:10px;background:rgba(0,229,160,.12);color:var(--gr);border:1px solid rgba(0,229,160,.2)}
.sb-lic-badge.warn{background:rgba(255,209,102,.12);color:var(--yw);border-color:rgba(255,209,102,.2)}
.sb-lic-badge.exp{background:rgba(255,79,107,.12);color:var(--rd);border-color:rgba(255,79,107,.2)}
.sb-lic-bar{height:3px;background:var(--b2);border-radius:2px;margin-top:5px;overflow:hidden}
.sb-lic-bar-fill{height:100%;border-radius:2px;background:var(--gr)}
.sb-footer{padding:8px;border-top:1px solid var(--b1);display:flex;gap:4px;flex-shrink:0}
.sb-btn{flex:1;padding:6px 2px;border-radius:7px;border:1px solid var(--b1);background:var(--s2);color:var(--t2);font-size:.62rem;cursor:pointer;font-family:var(--font);display:flex;align-items:center;justify-content:center;white-space:nowrap}
.sb-btn:hover{border-color:var(--ac);color:var(--ac)}

/* MAIN */
.main{flex:1;display:flex;flex-direction:column;overflow:hidden;min-width:0}
.main-header{padding:12px 18px;border-bottom:1px solid var(--b1);display:flex;align-items:center;justify-content:space-between;background:var(--s1);flex-shrink:0}
.main-title{font-size:.92rem;font-weight:700;display:flex;align-items:center;gap:8px}
.main-title .dot{width:7px;height:7px;border-radius:50%;background:var(--gr);flex-shrink:0}
.rc-badge{font-size:.75rem;color:var(--t2);background:var(--s3);padding:3px 10px;border-radius:20px;border:1px solid var(--b1)}
.main-content{flex:1;overflow-y:auto;padding:14px 18px}

/* MODAL BUTONLARI */
.mode-btns{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.mode-btn{padding:8px 18px;border-radius:8px;border:1.5px solid var(--b1);background:var(--s2);color:var(--t2);cursor:pointer;font-size:.82rem;font-weight:600;font-family:var(--font);transition:all .15s}
.mode-btn:hover{border-color:var(--b2);color:var(--t1)}
.mode-btn.active{border-color:var(--ac);background:var(--glow-ac);color:var(--ac)}

/* ARAMA FORMU */
.sf{background:var(--s1);border:1px solid var(--b1);border-radius:12px;padding:14px;margin-bottom:14px;display:none}
.sf.show{display:block}
.sf-row{display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end}
.fd{display:flex;flex-direction:column;gap:4px}
.fd label{font-size:.68rem;font-weight:700;color:var(--t2);text-transform:uppercase;letter-spacing:.07em}
.fd input,.fd select{padding:9px 12px;background:var(--s2);border:1.5px solid var(--b2);border-radius:7px;color:var(--t1);font-size:.85rem;outline:none;font-family:var(--font);min-width:130px}
.fd input:focus,.fd select:focus{border-color:var(--ac)}
.fd input.tc-f{min-width:170px;font-family:monospace;letter-spacing:.08em;font-size:.95rem}
.multi-row{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}
.multi-div{display:flex;align-items:center;gap:10px;margin:10px 0 8px;color:var(--t3);font-size:.7rem;font-weight:600;text-transform:uppercase;letter-spacing:.07em}
.multi-div::before,.multi-div::after{content:'';flex:1;height:1px;background:var(--b1)}
.btn-s{padding:9px 18px;background:var(--ac);border:none;border-radius:7px;color:#fff;font-weight:700;cursor:pointer;font-size:.82rem;font-family:var(--font);white-space:nowrap}
.btn-s:hover{opacity:.88}
.btn-c{padding:9px 12px;background:transparent;border:1.5px solid var(--b2);border-radius:7px;color:var(--t2);cursor:pointer;font-size:.82rem;font-family:var(--font)}
.btn-c:hover{color:var(--t1)}
.em{padding:9px 14px;background:rgba(255,79,107,.08);border:1px solid rgba(255,79,107,.2);border-radius:7px;color:var(--rd);font-size:.8rem;margin-bottom:10px;display:none}
.em.show{display:block}

/* RESULTS */
.rc{background:var(--s1);border:1px solid var(--b1);border-radius:12px;overflow:hidden;width:100%;box-sizing:border-box}
.rc-head{padding:11px 16px;border-bottom:1px solid var(--b1);display:flex;align-items:center;justify-content:space-between;gap:8px}
.rc-title{font-size:.82rem;color:var(--t2);font-weight:500;flex:1}
.badge-ok{padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:700;background:rgba(0,229,160,.12);color:var(--gr);border:1px solid rgba(0,229,160,.2)}
.badge-no{padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:700;background:var(--s3);color:var(--t3);border:1px solid var(--b1)}
.btn-excel{padding:5px 12px;background:rgba(0,229,160,.12);border:1px solid rgba(0,229,160,.2);color:var(--gr);border-radius:6px;cursor:pointer;font-size:.72rem;font-weight:600;font-family:var(--font)}
.btn-excel:hover{background:rgba(0,229,160,.2)}
.tw{overflow-x:auto;max-height:calc(100vh - 260px)}
table{width:100%;border-collapse:collapse;font-size:.8rem}
th{position:sticky;top:0;padding:9px 12px;text-align:left;font-size:.67rem;font-weight:700;color:var(--t2);background:var(--s2);border-bottom:1px solid var(--b1);white-space:nowrap;text-transform:uppercase;letter-spacing:.07em;z-index:1}
td{padding:8px 12px;border-bottom:1px solid var(--b1);color:var(--t1);white-space:nowrap}
tr:last-child td{border-bottom:none}
tr:hover td{background:var(--glow-ac)}
.es{padding:36px;text-align:center;color:var(--t3);font-size:.85rem}
.ld{padding:32px;text-align:center;color:var(--t2)}
.spin{display:inline-block;width:18px;height:18px;border:2px solid var(--b2);border-top-color:var(--ac);border-radius:50%;animation:sp .7s linear infinite;margin-right:8px;vertical-align:middle}
@keyframes sp{to{transform:rotate(360deg)}}

/* AİLE AĞACI */
.aile-tw{overflow-x:auto;overflow-y:auto;max-height:calc(100vh - 200px);width:100%}
.aile-tw table{width:100%;table-layout:auto;border-collapse:collapse}
.aile-tw th{padding:7px 10px;font-size:.78rem;font-family:var(--font);color:var(--t2);font-weight:600;background:var(--s2);border-bottom:1.5px solid var(--b1);white-space:nowrap;text-align:left}

.aile-tw td{white-space:nowrap;padding:7px 10px;font-size:.82rem;font-family:var(--font);color:var(--t1)}
.aile-tw td:nth-child(9){white-space:normal;word-break:break-all}
.rol-b{display:inline-block;padding:2px 8px;border-radius:5px;font-size:.72rem;font-weight:700;background:var(--s3);color:var(--t2);border:1px solid var(--b1);font-family:var(--font)}
.deg-btn{padding:6px 14px;border-radius:7px;border:1.5px solid var(--b1);background:var(--s2);color:var(--t2);cursor:pointer;font-size:.78rem;font-weight:600;font-family:var(--font);transition:all .15s}
.deg-btn:hover{border-color:var(--b2);color:var(--t1)}
.deg-btn.active{border-color:var(--ac);background:var(--glow-ac);color:var(--ac)}
.ar-merkez td{background:rgba(79,158,255,.07)!important}
.ar-merkez .rol-b{background:rgba(79,158,255,.15);color:var(--ac);border-color:rgba(79,158,255,.2)}
.ar-baba .rol-b{background:rgba(34,211,238,.1);color:var(--cy);border-color:rgba(34,211,238,.2)}
.ar-anne .rol-b{background:rgba(192,132,252,.1);color:var(--pr);border-color:rgba(192,132,252,.2)}
.ar-es .rol-b{background:rgba(244,114,182,.1);color:var(--pk);border-color:rgba(244,114,182,.2)}
.aile-sec td{padding:7px 12px;font-size:.65rem;font-weight:700;color:var(--t3);text-transform:uppercase;letter-spacing:.1em;background:var(--bg);border-top:1px solid var(--b1);border-bottom:1px solid var(--b1)}
.gsm-cell{color:var(--gr);font-family:monospace;font-size:.73rem}
.tree-btn{font-size:.67rem;padding:3px 8px;background:transparent;border:1px solid var(--b2);border-radius:5px;color:var(--t2);cursor:pointer;font-family:var(--font);white-space:nowrap}
.tree-btn:hover{border-color:var(--ac);color:var(--ac)}
.update-bar{padding:10px 16px;background:rgba(255,209,102,.08);border:1px solid rgba(255,209,102,.3);border-radius:8px;color:var(--yw);font-size:.82rem;margin-bottom:12px;display:none;flex-direction:column;gap:8px}
.update-bar.show{display:flex}
.update-bar .btn-upd{padding:5px 14px;background:var(--yw);border:none;border-radius:6px;color:#000;font-weight:700;cursor:pointer;font-size:.78rem;font-family:var(--font)}
.upd-progress{height:6px;background:var(--b2);border-radius:4px;overflow:hidden;display:none}
.upd-progress-fill{height:100%;background:var(--yw);border-radius:4px;transition:width .3s}
.info-bar{padding:8px 14px;background:rgba(79,158,255,.08);border:1px solid rgba(79,158,255,.2);border-radius:7px;font-size:.8rem;color:var(--ac);margin-bottom:10px;display:none}
.info-bar.show{display:block}
.upload-area{padding:16px;border:2px dashed var(--b2);border-radius:10px;text-align:center;cursor:pointer;transition:all .2s}
.upload-area:hover{border-color:var(--ac);background:var(--glow-ac)}
</style>
</head>
<body>
<div id="ls">
  <div class="lc">
    <div class="lc-top"><div class="lc-icon">🗂</div><h2>Sorgu Sistemi</h2><p>Güvenli erişim için kimlik doğrulama</p></div>
    <div class="lf"><label>Şifre</label><input type="password" id="pwd" placeholder="••••••••"></div>
    <button class="lbtn" onclick="doLogin()">Giriş Yap</button>
    <div class="lerr" id="lerr">Hatalı şifre.</div>
  </div>
</div>

<div id="app">
  <div class="sidebar">
    <div class="sb-header">
      <div class="sb-logo"><div class="sb-logo-icon">🗂</div><div><div>Sorgu Sistemi</div><div class="sb-sub">v4.0</div></div></div>
    </div>
    <div class="sb-nav" id="sb-nav"><div class="sb-sec">Sorgular</div></div>
    <div class="sb-lic" id="sb-lic" style="display:none">
      <div style="display:flex;align-items:center;justify-content:space-between">
        <span style="font-size:.67rem;color:var(--t2);font-weight:600">LİSANS</span>
        <span class="sb-lic-badge" id="lic-badge">—</span>
      </div>
      <div class="sb-lic-bar"><div class="sb-lic-bar-fill" id="lic-bar" style="width:0%"></div></div>
    </div>
    <div class="sb-footer">
      <button class="sb-btn" onclick="testDBs()" title="DB Test">🔌</button>
      <button class="sb-btn" onclick="toggleTheme()" id="theme-btn">🌙</button>
      <button class="sb-btn" onclick="doLogout()" style="flex:2">Çıkış →</button>
    </div>
  </div>

  <div class="main">
    <div class="main-header">
      <div class="main-title"><div class="dot"></div><span id="main-title">Sorgu Sistemi</span></div>
      <span class="rc-badge" id="rc-badge" style="display:none"></span>
    </div>
    <div class="main-content">
      <div class="update-bar" id="update-bar">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:10px">
          <span id="update-msg">🆕 Yeni sürüm mevcut</span>
          <button class="btn-upd" id="update-btn" onclick="doUpdate()">⬇️ Güncelle</button>
        </div>
        <div class="upd-progress" id="upd-progress">
          <div class="upd-progress-fill" id="upd-fill" style="width:0%"></div>
        </div>
        <div id="upd-warn" style="font-size:.75rem;color:var(--rd);display:none;font-weight:600">
          ⚠️ Güncelleme tamamlanana kadar uygulamayı KAPATMAYIN!
        </div>
        <div id="upd-pct" style="font-size:.75rem;text-align:center;display:none">%0 indiriliyor...</div>
      </div>
      <div class="em" id="em"></div>
      <div class="info-bar" id="info-bar"></div>

      <!-- ARAMA FORMU -->
      <div class="sf" id="sf">
        <!-- Mod butonları (çoklu mod için) -->
        <div class="mode-btns" id="mode-btns" style="display:none"></div>
        <!-- Ana TC/değer girişi -->
        <div class="sf-row" id="sf-main-row">
          <div class="fd" id="sf-main-field">
            <label id="sf-label">TC Kimlik No</label>
            <input class="tc-f" type="text" id="tc" placeholder="00000000000" maxlength="11" oninput="this.value=this.value.replace(/[^0-9]/g,'')">
          </div>
          <div class="fd" id="tc-son-wrap" style="display:none">
            <label>TC Son 2 Hane</label>
            <input type="text" id="tc-son" placeholder="örn: 60" maxlength="2" style="width:80px"
              oninput="this.value=this.value.replace(/[^0-9]/g,'')">
          </div>
          <div id="extra-fields" style="display:flex;gap:8px;flex-wrap:wrap"></div>
          <button class="btn-s" onclick="doSearch()">🔍 Ara</button>
          <button class="btn-c" onclick="doClear()">✕</button>
        </div>
        <!-- Aile derece seçimi -->
      <div id="aile-degree-bar" style="display:none;gap:6px;flex-wrap:wrap;margin-top:8px">
        <span style="font-size:.72rem;color:var(--t2);font-weight:600;align-self:center">Derece:</span>
        <button class="deg-btn active" onclick="setAileDegree(1,this)">1. Derece</button>
        <button class="deg-btn" onclick="setAileDegree(2,this)">2. Derece</button>
        <button class="deg-btn" onclick="setAileDegree(3,this)">3. Derece</button>
        <button class="deg-btn" onclick="setAileDegree(4,this)">4. Derece</button>
        <input type="hidden" id="aile-degree" value="1">
      </div>

      <!-- Çoklu alan (tcpro) -->
        <div id="multi-area" style="display:none">
          <div class="multi-div">veya isim / bilgi ile ara</div>
          <div class="multi-row">
            <div class="fd"><label>Ad</label><input type="text" id="m-AD" placeholder="Ad..."></div>
            <div class="fd"><label>Soyad</label><input type="text" id="m-SOYAD" placeholder="Soyad..."></div>
            <div class="fd"><label>Baba Adı</label><input type="text" id="m-BABAADI" placeholder="Baba Adı..."></div>
            <div class="fd"><label>Anne Adı</label><input type="text" id="m-ANNEADI" placeholder="Anne Adı..."></div>
          </div>
        </div>
      </div>

      <!-- TOPLU GSM yükleme alanı -->
      <div id="toplu-area" style="display:none;margin-top:4px">
        <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;padding:4px 0">
          <label for="toplu-file"
            style="padding:10px 20px;background:var(--s3);border:2px dashed var(--b2);border-radius:9px;cursor:pointer;font-size:.84rem;color:var(--t2);font-weight:600;transition:all .2s;display:flex;align-items:center;gap:8px"
            onmouseover="this.style.borderColor='var(--ac)';this.style.color='var(--ac)'"
            onmouseout="this.style.borderColor='var(--b2)';this.style.color='var(--t2)'">
            📂 Excel Seç <span style="font-size:.72rem;opacity:.7">(A sütununda TC listesi)</span>
          </label>
          <input type="file" id="toplu-file" accept=".xlsx,.xls" style="display:none" onchange="topluFileSelected(this)">
          <span id="toplu-file-info" style="font-size:.82rem;color:var(--t2)">Dosya seçilmedi</span>
          <button class="btn-s" id="toplu-btn" onclick="topluSorgula()" style="margin-left:8px">🔍 Sorgula & İndir</button>
          <button class="btn-c" onclick="topluTemizle()">✕ Temizle</button>
        </div>
        <div id="toplu-warn" style="display:none;margin-top:8px;padding:8px 12px;background:rgba(255,209,102,.08);border:1px solid rgba(255,209,102,.3);border-radius:7px;font-size:.78rem;color:var(--yw)">⚠️ 2. derece sorgu daha uzun sürebilir. Büyük TC listeleri için sabırla bekleyin.</div>
        <div id="toplu-tc-preview" style="margin-top:6px;font-size:.76rem;color:var(--t3);font-family:monospace"></div>
      </div>

      <!-- SONUÇLAR -->
      <div class="rc" id="rc" style="display:none">
        <div class="rc-head">
          <span class="rc-title" id="rt">—</span>
          <span id="rb" class="badge-no">0</span>
          <button class="btn-excel" onclick="exportExcel()" id="excel-btn" style="display:none">⬇️ Excel</button>
        </div>
        <div id="rb2"></div>
      </div>
    </div>
  </div>
</div>

<script>
let cfg={},adb=null,amode=null,lastRows=[],lastCols=[],_adminMod=false;
let isDark=true;

document.getElementById('pwd').onkeydown=e=>{if(e.key==='Enter')doLogin()};
document.addEventListener('keydown',e=>{
  if(e.key!=='Enter'||!document.getElementById('app').style.display||document.getElementById('app').style.display==='none')return;
  doSearch();
});

async function doLogin(){
  const r=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({password:document.getElementById('pwd').value})});
  if(r.ok){
    await loadCfg();
    document.getElementById('ls').style.display='none';
    document.getElementById('app').style.display='flex';
    setTimeout(checkUpdate, 2000);  // 2s sonra güncelleme kontrol
  } else {
    document.getElementById('lerr').style.display='block';
    document.getElementById('pwd').value='';document.getElementById('pwd').focus();
  }
}
async function doLogout(){await fetch('/api/logout',{method:'POST'});location.reload();}

function toggleTheme(){
  isDark=!isDark;
  document.documentElement.setAttribute('data-theme',isDark?'dark':'light');
  document.getElementById('theme-btn').textContent=isDark?'🌙':'☀️';
}

async function loadCfg(){
  const res=await(await fetch('/api/config')).json();
  // dbs_list: sıralı dizi, cfg: key→value haritası
  cfg={};
  const dbsList=res.dbs_list||[];
  dbsList.forEach(v=>cfg[v.key]=v);
  const nav=document.getElementById('sb-nav');
  nav.innerHTML='<div class="sb-sec">Sorgular</div>';
  dbsList.forEach(v=>{const k=v.key;
    const el=document.createElement('div');
    el.className='nav-item';el.dataset.key=k;
    const icon=document.createElement('div');
    icon.className='nav-icon';
    icon.style.cssText=`background:${v.color}22;border-color:${v.color}33;color:${v.color}`;
    icon.textContent=v.icon;
    const label=document.createElement('span');
    label.className='nav-label';label.textContent=v.label;
    el.appendChild(icon);el.appendChild(label);
    el.onclick=()=>selDb(k);
    nav.appendChild(el);
  }); // end dbsList.forEach
  // Admin mod global
  _adminMod = res.license?.admin_mod || false;

  // Lisans
  const lic=res.license;
  if(lic){
    document.getElementById('sb-lic').style.display='block';
    const k=lic.kalan;
    const badge=document.getElementById('lic-badge');
    const bar=document.getElementById('lic-bar');
    if(lic.end_date==='SINIRSIZ'||k===null){
      badge.textContent='∞ Sınırsız';badge.className='sb-lic-badge';
      bar.style.width='100%';bar.style.background='var(--gr)';
    } else if(k<=0){
      badge.textContent='Süresi Doldu';badge.className='sb-lic-badge exp';bar.style.width='0%';
    } else {
      badge.textContent=k<=30?`⚠️ ${k} gün`:`${k} gün`;
      badge.className='sb-lic-badge'+(k<=30?' warn':'');
      const total=lic.plan==='AYLIK'?30:lic.plan==='YILLIK'?365:k;
      bar.style.width=Math.min(100,Math.round((k/total)*100))+'%';
      bar.style.background=k<=30?'var(--yw)':'var(--gr)';
    }
  }
}


async function checkUpdate(){
  try{
    const r=await fetch('/api/update/check');
    const d=await r.json();
    if(d.has_update){
      const bar=document.getElementById('update-bar');
      document.getElementById('update-msg').textContent=
        `🆕 Yeni sürüm: ${d.remote} (${d.date}) — ${d.notes||''}`;
      bar.classList.add('show');
    }
  }catch(e){}
}

let _updateInterval=null;
async function doUpdate(){
  if(!confirm('Güncelleme indirilecek. Tamamlanana kadar kapatmayın. Devam?'))return;
  document.getElementById('update-btn').disabled=true;
  document.getElementById('update-btn').textContent='⏳ Başlatılıyor...';
  document.getElementById('upd-progress').style.display='block';
  document.getElementById('upd-warn').style.display='block';
  document.getElementById('upd-pct').style.display='block';
  try{
    const r=await fetch('/api/update/download',{method:'POST'});
    const d=await r.json();
    if(!r.ok){alert('Hata: '+(d.error||''));resetUpdateUI();return;}
    // Progress takibi
    _updateInterval=setInterval(async()=>{
      try{
        const p=await(await fetch('/api/update/progress')).json();
        document.getElementById('upd-fill').style.width=p.pct+'%';
        document.getElementById('upd-pct').textContent='%'+p.pct+' indiriliyor...';
        document.getElementById('update-msg').textContent='⬇️ Güncelleme indiriliyor — %'+p.pct;
        if(p.status==='done'){
          clearInterval(_updateInterval);
          document.getElementById('upd-pct').textContent='✅ İndirme tamamlandı!';
          document.getElementById('upd-warn').style.display='none';
          document.getElementById('update-msg').textContent='✅ Güncelleme hazır!';
          document.getElementById('upd-fill').style.background='var(--gr)';
          document.getElementById('update-btn').textContent='🔄 Yeniden Başlat';
          document.getElementById('update-btn').disabled=false;
          document.getElementById('update-btn').onclick=async()=>{
            if(!confirm('Uygulama kapanıp yeniden başlayacak. Devam?'))return;
            await fetch('/api/update/apply',{method:'POST'});
            setTimeout(()=>{ window.location.reload(); },8000);
          };
        } else if(p.status==='error'){
          clearInterval(_updateInterval);
          alert('İndirme hatası: '+p.error);
          resetUpdateUI();
        }
      }catch(e){}
    },1000);
  }catch(e){alert('Hata: '+e.message);resetUpdateUI();}
}
function resetUpdateUI(){
  document.getElementById('update-btn').disabled=false;
  document.getElementById('update-btn').textContent='⬇️ Güncelle';
  document.getElementById('upd-progress').style.display='none';
  document.getElementById('upd-warn').style.display='none';
  document.getElementById('upd-pct').style.display='none';
  if(_updateInterval)clearInterval(_updateInterval);
}

function selDb(k){
  // Derece butonlarını gizle (aile dışında)
  const degBar=document.getElementById('aile-degree-bar');
  if(degBar) degBar.style.display=(k==='aile')?'flex':'none';
  // TC Son 2 Hane sadece tcpro'da görünsün
  const tcSonWrap=document.getElementById('tc-son-wrap');
  if(tcSonWrap){tcSonWrap.style.display=(k==='tcpro')?'flex':'none';}
  const tcSon=document.getElementById('tc-son');
  if(tcSon&&k!=='tcpro')tcSon.value='';

  // Toplu GSM - admin_mod kontrolü
  if(k==='toplu'){
    document.querySelectorAll('.nav-item').forEach(n=>n.classList.toggle('active',n.dataset.key===k));
    document.getElementById('main-title').textContent='📋 Toplu GSM Sorgu';
    hideErr();hideInfo();
    if(!_adminMod){
      // Kilitli
      document.getElementById('sf').classList.remove('show');
      document.getElementById('rc').style.display='block';
      document.getElementById('rb').textContent='Kilitli';
      document.getElementById('rb').className='badge-no';
      document.getElementById('rc-badge').style.display='none';
      document.getElementById('excel-btn').style.display='none';
      document.getElementById('rb2').innerHTML=`
        <div style="padding:48px;text-align:center">
          <div style="font-size:2.5rem;margin-bottom:12px">🔒</div>
          <div style="font-size:1rem;font-weight:700;color:var(--t1);margin-bottom:8px">Bu sorgulama ücretlidir.</div>
          <div style="font-size:.85rem;color:var(--t2)">Yönetici ile iletişime geçin.</div>
        </div>`;
      return;
    }
    // Admin mod - normal selDb akışına devam et
  }
  adb=k;amode=null;
  document.querySelectorAll('.nav-item').forEach(n=>n.classList.toggle('active',n.dataset.key===k));
  const v=cfg[k];
  document.getElementById('main-title').textContent=`${v.icon} ${v.label}`;
  document.getElementById('rc').style.display='none';
  document.getElementById('rc-badge').style.display='none';
  document.getElementById('sf').classList.add('show');
  hideErr();hideInfo();
  document.getElementById('multi-area').style.display=v.multi_search?'block':'none';
  document.getElementById('extra-fields').innerHTML='';

  // Mod butonları
  const mb=document.getElementById('mode-btns');
  if(v.multi_modal&&v.modes&&v.modes.length){
    mb.style.display='flex';mb.innerHTML='';
    v.modes.forEach((m,i)=>{
      const b=document.createElement('button');
      b.className='mode-btn'+(i===0?' active':'');
      b.textContent=m.label;b.dataset.mode=m.key;
      b.onclick=()=>selMode(k,m.key,b);
      mb.appendChild(b);
    });
    selMode(k,v.modes[0].key,mb.children[0]);
  } else {
    mb.style.display='none';
    setupDefaultForm(k);
  }
  setTimeout(()=>document.getElementById('tc').focus(),50);
}

function selMode(dbKey,modeKey,btn){
  amode=modeKey;
  document.querySelectorAll('.mode-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  setupModeForm(dbKey,modeKey);
}

function setupDefaultForm(k){
  const tc=document.getElementById('tc');
  tc.type='text';
  tc.oninput=function(){this.value=this.value.replace(/[^0-9]/g,'')};
  if(k==='tcpro'){
    tc.placeholder='TC (tam 11 hane veya ilk 3-9 hane)';
    tc.maxLength=11;
    tc.removeAttribute('required');
    document.getElementById('tc-son-wrap').style.display='block';
  } else {
    document.getElementById('tc-son-wrap').style.display='none';
    document.getElementById('tc-son').value='';
    tc.placeholder='00000000000';
    tc.maxLength=11;
  }
  document.getElementById('sf-label').textContent='TC Kimlik No';
  // Filtre alanları
  const v=cfg[k];
  const ef=document.getElementById('extra-fields');
  ef.innerHTML='';
  if(v&&v.filters&&v.filters.length&&k==='tcpro'){
    v.filters.forEach(f=>{
      const div=document.createElement('div');
      div.className='fd';
      div.innerHTML=`<label>${f.label}</label><input type="text" id="filter-${f.key}" placeholder="${f.label}..." style="min-width:110px">`;
      ef.appendChild(div);
    });
  }
  document.getElementById('toplu-area').style.display='none';
  document.getElementById('sf-main-row').style.display='flex';
  document.querySelector('.btn-s').onclick=doSearch;
  document.querySelector('.btn-s').textContent='🔍 Ara';
}

function setupModeForm(dbKey,modeKey){
  const tc=document.getElementById('tc');
  const ef=document.getElementById('extra-fields');
  ef.innerHTML='';
  document.getElementById('rc').style.display='none';

  // Toplu GSM modülü
  if(dbKey==='toplu'){
    document.getElementById('sf-main-row').style.display='none';
    document.getElementById('multi-area').style.display='none';
    document.getElementById('toplu-area').style.display='block';
    document.querySelector('.btn-s').onclick=topluSorgula;
    document.querySelector('.btn-s').textContent='🔍 Sorgula & İndir';
    // 2. derece uyarısı
    if(modeKey==='aile2'){
      document.getElementById('toplu-warn').style.display='block';
    } else {
      document.getElementById('toplu-warn').style.display='none';
    }
    // Dosya sıfırla
    topluTcList=[];
    document.getElementById('toplu-file-info').textContent='Dosya seçilmedi';
    document.getElementById('toplu-tc-preview').textContent='';
    document.getElementById('toplu-file').value='';
    return;
  }
  document.getElementById('toplu-area').style.display='none';
  document.getElementById('sf-main-row').style.display='flex';
  document.querySelector('.btn-s').onclick=doSearch;
  document.querySelector('.btn-s').textContent='🔍 Ara';

  if(dbKey==='gsm'){
    if(modeKey==='tc2gsm'||modeKey==='tc2vergi'){
      document.getElementById('sf-label').textContent='TC Kimlik No';
      tc.placeholder='00000000000';tc.maxLength=11;
      tc.oninput=function(){this.value=this.value.replace(/[^0-9]/g,'')};
    } else if(modeKey==='gsm2tc'){
      document.getElementById('sf-label').textContent='GSM Numarası';
      tc.placeholder='05xxxxxxxxx';tc.maxLength=11;
      tc.oninput=function(){this.value=this.value.replace(/[^0-9]/g,'')};
    } else if(modeKey==='vergi2tc'){
      document.getElementById('sf-label').textContent='Vergi Numarası';
      tc.placeholder='Vergi No...';tc.maxLength=20;
      tc.oninput=null;
    }
  } else if(dbKey==='tapu'){
    if(modeKey==='tc2tapu'){
      document.getElementById('sf-label').textContent='TC Kimlik No';
      tc.placeholder='00000000000';tc.maxLength=11;
      tc.oninput=function(){this.value=this.value.replace(/[^0-9]/g,'')};
    } else if(modeKey==='parsel2kisi'){
      document.getElementById('sf-label').textContent='İl';
      tc.placeholder='İl adı...';tc.maxLength=50;tc.oninput=null;
      ef.innerHTML=`
        <div class="fd"><label>İlçe</label><input type="text" id="f-ilce" placeholder="İlçe..."></div>
        <div class="fd"><label>Ada</label><input type="text" id="f-ada" placeholder="Ada No..."></div>
        <div class="fd"><label>Parsel</label><input type="text" id="f-parsel" placeholder="Parsel No..."></div>`;
    }
  } else if(dbKey==='sgk'){
    document.getElementById('sf-label').textContent='TC Kimlik No';
    tc.placeholder='00000000000';tc.maxLength=11;
    tc.oninput=function(){this.value=this.value.replace(/[^0-9]/g,'')};
  }
  tc.value='';
}

async function doSearch(){
  if(!adb)return;
  hideErr();hideInfo();
  const tc=document.getElementById('tc').value.trim();
  const v=cfg[adb];
  // Filtreler
  const filters={};
  if(v&&v.filters){v.filters.forEach(f=>{const el=document.getElementById('filter-'+f.key);if(el)filters[f.key]=el.value.trim();});}

  // Özel endpoint'ler
  if(adb==='gsm'&&amode){
    if(!tc){showErr('Değer giriniz');return;}
    showLoading();
    await fetchAndRender('/api/gsm_query',{mode:amode,val:tc},'gsm');
    return;
  }
  if(adb==='tapu'&&amode){
    if(amode==='tc2tapu'){
      if(!tc||tc.length!==11){showErr('11 haneli TC giriniz');return;}
      showLoading();
      await fetchAndRender('/api/tapu_query',{mode:'tc2tapu',tc},'tapu');
    } else {
      const il=tc;
      const ilce=document.getElementById('f-ilce')?.value||'';
      const ada=document.getElementById('f-ada')?.value||'';
      const parsel=document.getElementById('f-parsel')?.value||'';
      if(!il&&!ilce&&!ada&&!parsel){showErr('En az bir kriter giriniz');return;}
      showLoading();
      await fetchAndRender('/api/tapu_query',{mode:'parsel2kisi',il,ilce,ada,parsel},'tapu');
    }
    return;
  }
  if(adb==='sgk'&&amode){
    if(!tc||tc.length!==11){showErr('11 haneli TC giriniz');return;}
    showLoading();
    await fetchAndRender('/api/sgk_query',{mode:amode,tc},'sgk');
    return;
  }
  if(adb==='komsular'){
    if(!tc||tc.length!==11){showErr('11 haneli TC giriniz');return;}
    showLoading();
    try{
      const r=await fetch('/api/komsular',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tc})});
      const d=await r.json();
      if(!r.ok){showErr(d.error||'Hata');document.getElementById('rc').style.display='none';return;}
      if(d.adres) showInfo(`Adres: ${d.adres}`);
      renderTable(d,tc);
    }catch(e){showErr('Bağlantı hatası: '+e.message);}
    return;
  }
  if(adb==='aile'){
    if(!tc||tc.length!==11){showErr('11 haneli TC giriniz');return;}
    const deg=parseInt(document.getElementById('aile-degree')?.value||'1');
    showLoading();
    try{
      const r=await fetch('/api/aile',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tc,degree:deg})});
      const d=await r.json();
      if(!r.ok||!d.kisi){document.getElementById('rb2').innerHTML='<div class="es">Bu TC için kayıt bulunamadı.</div>';return;}
      renderAile(d,tc);
    }catch(e){showErr('Bağlantı hatası: '+e.message);}
    return;
  }
  // Normal DB sorgusu
  const multi={};
  ['AD','SOYAD','BABAADI','ANNEADI','DOGUMTARIHI','DOGUMYERI','MEMLEKETIL','MEMLEKETILCE'].forEach(k=>{
    const el=document.getElementById('m-'+k);if(el)multi[k]=el.value.trim();
  });
  const hasMulti=Object.values(multi).some(v=>v);
  const hasFilter=Object.values(filters).some(v=>v);
  if(!tc&&!hasMulti&&!hasFilter){showErr(v.multi_search?'TC veya en az bir kriter giriniz':'TC giriniz');return;}
  if(tc&&!/^\d+$/.test(tc)){showErr('TC sadece rakam içerebilir');return;}
  if(tc&&tc.length!==11&&adb!=='tcpro'){showErr('11 haneli TC giriniz');return;}
  showLoading();
  try{
    const r=await fetch('/api/query',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({db:adb,tc,tc_son:document.getElementById('tc-son')?.value.trim()||'',filters,multi})});
    let d;try{d=await r.json();}catch(e){showErr('Yanıt okunamadı');return;}
    if(!r.ok){showErr(d.error||`Hata (${r.status})`);document.getElementById('rc').style.display='none';return;}
    renderTable(d,tc||'çoklu arama');
  }catch(e){showErr('Bağlantı hatası: '+e.message);}
}

async function fetchAndRender(url,body,tag){
  try{
    const r=await fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
    let d;try{d=await r.json();}catch(e){showErr('Yanıt okunamadı');return;}
    if(!r.ok){showErr(d.error||'Hata');document.getElementById('rc').style.display='none';return;}
    if(d.info) showInfo(d.info);
    renderTable(d,body.tc||body.val||'');
  }catch(e){showErr('Bağlantı hatası: '+e.message);}
}

function renderTable(d,label){
  lastRows=d.rows||[];lastCols=d.columns||[];
  const rc=document.getElementById('rc');rc.style.display='block';
  const rb=document.getElementById('rb');
  rb.textContent=`${d.count} kayıt`;rb.className=d.count?'badge-ok':'badge-no';
  document.getElementById('rc-badge').style.display='block';
  document.getElementById('rc-badge').textContent=`${d.count} kayıt`;
  document.getElementById('rt').textContent=`${cfg[adb]?.icon||''} ${cfg[adb]?.label||''} — ${label}`;
  document.getElementById('excel-btn').style.display=d.count?'block':'none';
  if(!d.rows.length){document.getElementById('rb2').innerHTML='<div class="es">Kayıt bulunamadı.</div>';return;}
  let h='<div class="tw"><table><thead><tr>';
  d.columns.forEach(c=>h+=`<th>${c}</th>`);
  h+='</tr></thead><tbody>';
  d.rows.forEach(row=>{h+='<tr>';d.columns.forEach(c=>h+=`<td>${esc(row[c]??'—')}</td>`);h+='</tr>';});
  h+='</tbody></table></div>';
  document.getElementById('rb2').innerHTML=h;
}

function exportExcel(){
  if(!lastRows.length)return;
  // SheetJS henüz yüklenmediyse yükle
  if(typeof XLSX==='undefined'){
    const s=document.createElement('script');
    s.src='https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js';
    s.onload=()=>doExport();
    s.onerror=()=>exportCSV();  // internet yoksa CSV
    document.head.appendChild(s);
    return;
  }
  doExport();
}
function doExport(){
  if(!lastRows.length)return;
  // GSM sütunu varsa birden fazla numara alt alta ayrı satır olur
  const gsmCols=['GSM','gsm'];
  const expanded=[];
  lastRows.forEach(r=>{
    // Hangi kolon GSM?
    let gsmCol=lastCols.find(c=>gsmCols.includes(c.toLowerCase()));
    if(!gsmCol){ // GSM kolonu yok - direkt ekle
      const obj={};lastCols.forEach(c=>obj[c]=r[c]??'');expanded.push(obj);return;
    }
    const gsmVal=String(r[gsmCol]||'—');
    const nums=gsmVal.split(', ').map(s=>s.trim()).filter(s=>s&&s!=='—');
    if(nums.length<=1){
      const obj={};lastCols.forEach(c=>obj[c]=r[c]??'');expanded.push(obj);
    } else {
      nums.forEach(gsm=>{
        const obj={};
        lastCols.forEach(c=>{obj[c]=c===gsmCol?gsm:(r[c]??'');});
        expanded.push(obj);
      });
    }
  });
  const ws=XLSX.utils.json_to_sheet(expanded,{header:lastCols});
  const wb=XLSX.utils.book_new();XLSX.utils.book_append_sheet(wb,ws,"Sonuçlar");
  const ts=new Date().toISOString().slice(0,19).replace(/:/g,'-');
  XLSX.writeFile(wb,`Sorgu_${cfg[adb]?.label||adb}_${ts}.xlsx`);
}
function exportCSV(){
  if(!lastRows.length)return;
  const rows=[lastCols.join(',')];
  lastRows.forEach(r=>rows.push(lastCols.map(c=>JSON.stringify(r[c]??'')).join(',')));
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(rows.join('\n'));
  a.download=`Sorgu_${cfg[adb]?.label||adb}.csv`;a.click();
}

function kisiRow(p,rol,cls=''){
  if(!p)return'';
  const ad=`${p.AD||''} ${p.SOYAD||''}`.trim()||'—';
  const tc=p.TC||'',dog=p.DOGUMTARIHI||'—';
  const il=p.MEMLEKETIL||p.ADRESIL||'—';
  const gsm=(p.__gsm||[]).map(g=>`<div>${g}</div>`).join('')||'—';
  const gsmTxt=(p.__gsm||[]).join(', ')||'—';
  lastRows.push({'Yakınlık':rol,'Ad Soyad':ad,'TC':tc,'Doğum':dog,'İl':il,'GSM':gsmTxt});
  return`<tr class="aile-row ${cls?'ar-'+cls:''}">
    <td><span class="rol-b">${rol}</span></td>
    <td>${ad}</td>
    <td><span style="font-family:monospace;font-size:.75rem">${tc}</span></td>
    <td>${dog}</td><td>${il}</td>
    <td class="gsm-cell">${gsm}</td>
    <td><button class="tree-btn" onclick="araAile('${tc}')">Ağaç →</button></td>
  </tr>`;
}
function secRow(b,s){if(!s.trim())return'';return`<tr class="aile-sec"><td colspan="9">${b}</td></tr>${s}`;}

function renderAile(d,tc){
  lastRows=[];
  lastCols=['Yakınlık','Ad Soyad','TC','Doğum','İl','GSM'];
  const deg=d.degree||1;
  const {kisi,derece1,derece2,derece3,es_tarafi}=d;
  const torunlar=d.torunlar||[];
  const yegenler=d.yegenler||[];
  const buyuk_torun=d.buyuk_torun||[];
  const kuzenler=d.kuzenler||[];
  const buyuk_amca=d.buyuk_amca||[];
  const buyuk_yegen=d.buyuk_yegen||[];
  const thead=`<thead><tr><th>Yakınlık</th><th>Ad Soyad</th><th>TC</th><th>Doğum</th><th>İl</th><th>GSM</th><th></th></tr></thead>`;
  let tb='';
  // Her zaman: Bizzat (★)
  tb+=secRow('⭐ Sorgulanan Kişi',kisiRow(kisi,'★ Bizzat','merkez'));

  // Sadece seçilen derece
  if(deg===1){
    let d1H=kisiRow(derece1.baba,'Baba','baba')+kisiRow(derece1.anne,'Anne','anne');
    derece1.cocuklar.forEach(c=>d1H+=kisiRow(c,'Çocuk'));
    if(d1H)tb+=secRow('1️⃣ 1. Derece — Anne / Baba / Çocuklar',d1H);
  } else if(deg===2){
    let d2H='';
    derece1.kardesler.forEach(k=>d2H+=kisiRow(k,'Kardeş'));
    d2H+=kisiRow(derece2.baba_baba,'Büyükbaba (B)')+kisiRow(derece2.baba_anne,'Büyükanne (B)');
    d2H+=kisiRow(derece2.anne_baba,'Büyükbaba (A)')+kisiRow(derece2.anne_anne,'Büyükanne (A)');
    torunlar.forEach(t=>d2H+=kisiRow(t,'Torun'));
    if(d2H)tb+=secRow('2️⃣ 2. Derece — Kardeşler / Büyükanne & Büyükbaba / Torunlar',d2H);
  } else if(deg===3){
    let d3H='';
    derece2.amca_hala.forEach(a=>d3H+=kisiRow(a,a.__rol||'Amca/Hala'));
    derece2.dayi_teyze.forEach(a=>d3H+=kisiRow(a,a.__rol||'Dayı/Teyze'));
    yegenler.forEach(y=>d3H+=kisiRow(y,'Yeğen'));
    buyuk_torun.forEach(b=>d3H+=kisiRow(b,'Büyük Torun'));
    if(d3H)tb+=secRow('3️⃣ 3. Derece — Amca / Dayı / Hala / Teyze / Yeğen / Büyük Torun',d3H);
  } else if(deg===4){
    let d4H='';
    kuzenler.forEach(k=>d4H+=kisiRow(k,'Kuzen'));
    buyuk_amca.forEach(b=>d4H+=kisiRow(b,b.__rol||'B. Amca/Dayı/Hala/Teyze'));
    buyuk_yegen.forEach(b=>d4H+=kisiRow(b,'Büyük Yeğen'));
    if(d4H)tb+=secRow('4️⃣ 4. Derece — Kuzenler / B. Amca & Teyze / Büyük Yeğen',d4H);
  }

  // Eş tarafı (her zaman göster)
  if(deg===1&&es_tarafi&&Object.keys(es_tarafi).length){
    let esT='';
    Object.values(es_tarafi).forEach((et,i)=>{
      const s=Object.keys(es_tarafi).length>1?` ${i+1}`:'';
      esT+=kisiRow(et.es,`Eş${s}`,'es');
      esT+=kisiRow(et.kayin_baba,`Kayınpeder${s}`)+kisiRow(et.kayin_anne,`Kayınvalide${s}`);
      (et.es_kardesler||[]).forEach(k=>esT+=kisiRow(k,'Kayın'));
    });
    if(esT)tb+=secRow('💍 Eş Tarafı',esT);
  }
  document.getElementById('rb').textContent='Aile ağacı';document.getElementById('rb').className='badge-ok';
  document.getElementById('excel-btn').style.display='block';
  document.getElementById('rt').textContent=`🧬 Bizzat ve Aile Sorgusu — TC: ${tc}`;
  document.getElementById('rc').style.display='block';
  document.getElementById('rb2').innerHTML=`<div class="aile-tw"><table>${thead}<tbody>${tb}</tbody></table></div>`;
}

function araAile(tc){if(!tc)return;selDb('aile');document.getElementById('tc').value=tc;doSearch();}
function doClear(){
  document.getElementById('tc').value='';
  const tcSon=document.getElementById('tc-son');if(tcSon)tcSon.value='';
  document.querySelectorAll('[id^="m-"],[id^="f-"]').forEach(e=>e.value='');
  document.getElementById('rc').style.display='none';
  document.getElementById('rc-badge').style.display='none';
  hideErr();hideInfo();
}
function showLoading(){
  const rc=document.getElementById('rc');rc.style.display='block';
  document.getElementById('rb2').innerHTML='<div class="ld"><span class="spin"></span>Sorgulanıyor...</div>';
  document.getElementById('rb').textContent='...';document.getElementById('rb').className='badge-no';
  document.getElementById('excel-btn').style.display='none';
}

// ── TOPLU GSM ────────────────────────────────────────────────────────────────
let topluTcList = [];

function topluFileSelected(input){
  const file = input.files[0];
  if(!file) return;
  document.getElementById('toplu-file-info').textContent = '⏳ Okunuyor...';
  // SheetJS ile oku
  loadSheetJS(()=>{
    const reader = new FileReader();
    reader.onload = (e)=>{
      try{
        const wb = XLSX.read(e.target.result, {type:'binary'});
        const ws = wb.Sheets[wb.SheetNames[0]];
        const data = XLSX.utils.sheet_to_json(ws, {header:1, raw:false});
        // İlk kolondan TC'leri al (sayısal, 11 hane)
        topluTcList = data
          .map(r => String(r[0]||'').trim().replace(/[^0-9]/g,''))
          .filter(t => t.length===11);
        document.getElementById('toplu-file-info').textContent =
          `✅ ${file.name} — ${topluTcList.length} TC bulundu`;
        document.getElementById('toplu-tc-preview').textContent =
          topluTcList.length>5
            ? `İlk 5: ${topluTcList.slice(0,5).join(', ')}...`
            : topluTcList.join(', ');
      }catch(err){
        document.getElementById('toplu-file-info').textContent = '❌ Dosya okunamadı: '+err.message;
      }
    };
    reader.readAsBinaryString(file);
  });
}

function loadSheetJS(cb){
  if(typeof XLSX!=='undefined'){cb();return;}
  const s=document.createElement('script');
  s.src='https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js';
  s.onload=cb;
  s.onerror=()=>alert('Excel kütüphanesi yüklenemedi. İnternet bağlantısını kontrol edin.');
  document.head.appendChild(s);
}

async function topluSorgula(){
  if(!topluTcList.length){showErr('Önce Excel dosyası seçin');return;}
  if(!amode){showErr('Mod seçin (Bizzat GSM / Toplu Aile GSM)');return;}
  hideErr();
  showLoading();
  document.getElementById('rb2').innerHTML=
    `<div class="ld"><span class="spin"></span>${topluTcList.length} TC sorgulanıyor...</div>`;

  try{
    const r=await fetch('/api/toplu_gsm_preview',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({mode:amode,tc_list:topluTcList})
    });
    let d;try{d=await r.json();}catch(e){showErr('Yanıt okunamadı');return;}
    if(!r.ok){showErr(d.error||'Hata');document.getElementById('rc').style.display='none';return;}

    // Sonuçları göster
    const rc=document.getElementById('rc');rc.style.display='block';
    document.getElementById('rb').textContent=`${d.rows?.length||0} satır`;
    document.getElementById('rb').className='badge-ok';
    document.getElementById('rt').textContent=`📋 Toplu GSM — ${topluTcList.length} TC`;

    // Tablo render
    const rows=d.rows||[];const cols=d.columns||[];
    let h='<div class="tw"><table><thead><tr>';
    cols.forEach(c=>h+=`<th>${c}</th>`);
    h+='</tr></thead><tbody>';
    rows.forEach(row=>{
      if(row.__sep){
        // Siyah ayırıcı satır
        h+=`<tr style="height:6px;background:#000"><td colspan="${cols.length}" style="background:#000;padding:0"></td></tr>`;
      } else {
        h+='<tr>';
        cols.forEach(c=>h+=`<td>${esc(row[c]??'—')}</td>`);
        h+='</tr>';
      }
    });
    h+='</tbody></table></div>';
    document.getElementById('rb2').innerHTML=h;

    // Excel indir butonu
    document.getElementById('excel-btn').style.display='block';
    document.getElementById('excel-btn').onclick=()=>topluExcelIndir();
    document.getElementById('excel-btn').textContent='⬇️ Excel İndir';

  }catch(e){
    showErr('Bağlantı hatası: '+e.message);
  }
}

async function topluExcelIndir(){
  if(!topluTcList.length)return;
  const btn=document.getElementById('excel-btn');
  btn.textContent='⏳ Hazırlanıyor...';btn.disabled=true;
  try{
    const r=await fetch('/api/toplu_gsm_excel',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({mode:amode,tc_list:topluTcList})
    });
    const ct=r.headers.get('Content-Type')||'';
    if(!r.ok||ct.includes('text/html')){
      // HTML hata sayfası geldi - otur hatayı text olarak oku
      const txt=await r.text();
      showErr('Sunucu hatası ('+r.status+'). Lütfen tekrar deneyin.');
      console.error('Excel hata:', txt.substring(0,200));
      return;
    }
    if(ct.includes('application/json')){
      const e=await r.json();showErr(e.error||'Hata');return;
    }
    const blob=await r.blob();
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url;
    a.download=`Toplu_GSM_${amode}_${new Date().toISOString().slice(0,10)}.xlsx`;
    a.click();URL.revokeObjectURL(url);
  }catch(e){
    showErr('İndirme hatası: '+e.message);
  } finally {
    btn.textContent='⬇️ Excel İndir';btn.disabled=false;
  }
}

function setAileDegree(deg,btn){
  document.getElementById('aile-degree').value=deg;
  document.querySelectorAll('.deg-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
}

function topluTemizle(){
  topluTcList=[];
  document.getElementById('toplu-file-info').textContent='Dosya seçilmedi';
  document.getElementById('toplu-tc-preview').textContent='';
  document.getElementById('toplu-file').value='';
  document.getElementById('rc').style.display='none';
  hideErr();
}

function showErr(m){console.error('[Sorgu]',m);const e=document.getElementById('em');e.textContent=m;e.style.display='block';e.classList.add('show');}
function hideErr(){const e=document.getElementById('em');e.style.display='none';e.classList.remove('show');}
function showInfo(m){const e=document.getElementById('info-bar');e.textContent=m;e.classList.add('show');}
function hideInfo(){document.getElementById('info-bar').classList.remove('show');}
function esc(s){return String(s??'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
async function testDBs(){
  showErr('🔌 Test ediliyor...');
  try{
    const r=await fetch('/api/test_db');const d=await r.json();
    const lines=Object.entries(d).map(([k,v])=>v.ok?`✅ ${k}: ${v.count?.toLocaleString()} kayıt`:`❌ ${k}: ${v.error}`).join('\n');
    alert('DB Test:\n\n'+lines);hideErr();
  }catch(e){showErr('Test hatası: '+e.message);}
}
</script>
</body>
</html>"""

if __name__=="__main__":
    def pf(p):
        with socket.socket(socket.AF_INET,socket.SOCK_STREAM) as s:
            return s.connect_ex(("localhost",p))!=0
    if not pf(5001): webbrowser.open("http://localhost:5001")
    else:
        log.info("Sorgu Sistemi v4.0 → http://localhost:5001")
        threading.Timer(1.5,lambda:webbrowser.open("http://localhost:5001")).start()
        app.run(host="127.0.0.1",port=5001,debug=False,threaded=True,use_reloader=False)
